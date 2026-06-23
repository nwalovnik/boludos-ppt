"""
Scraper diario de indicadores macro argentinos.

Carga el bedrock histórico (historical.json) y sobreescribe los registros recientes
con datos frescos de las APIs públicas. Output: data.json (en raíz del repo).

Fuentes:
  - INDEC (apis.datos.gob.ar/series)        → IPC, EMAE, IPI, ISAC, ICA, EPH, RIPTE, recaudación, BM
  - BCRA  (api.bcra.gob.ar/estadisticas/v4) → Reservas, TC oficial/mayorista, BM diaria, var IPC
  - ArgentinaDatos                          → EMBI riesgo país, cotizaciones (blue/mep/ccl)
  - Bluelytics                              → TC blue actual (respaldo)

Uso:
    python scrape.py            # corre todo, guarda data.json
    python scrape.py --dry-run  # corre todo pero no guarda

Diseño: idempotente, no rompe si una fuente falla (loggea y sigue).
"""
from __future__ import annotations

import argparse
import io
import json
import sys
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any

import requests
import xlrd  # noqa: E402  (.xls 97-2003 lectura)

# ──────────────────────────────────────────────────────────────────────────────
# Paths
# ──────────────────────────────────────────────────────────────────────────────
SCRAPER_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRAPER_DIR.parent
HISTORICAL = SCRAPER_DIR / "historical.json"
OUT_PATH = REPO_ROOT / "data.json"

# ──────────────────────────────────────────────────────────────────────────────
# Endpoints
# ──────────────────────────────────────────────────────────────────────────────
INDEC_API = "https://apis.datos.gob.ar/series/api/series/"
BCRA_API = "https://api.bcra.gob.ar/estadisticas/v4.0/monetarias/"
ARG_DATOS = "https://api.argentinadatos.com/v1"
BLUELYTICS = "https://api.bluelytics.com.ar/v2/latest"

# IDs INDEC (mismas que el HTML original + nuevas)
INDEC_SERIES = {
    "ipc":       "148.3_INIVELNAL_DICI_M_26",
    "emae_orig": "143.3_NO_PR_2004_A_21",
    "emae_dest": "143.3_NO_PR_2004_A_31",
    "ipi_orig":  "453.1_SERIE_ORIGNAL_0_0_14_46",
    "ipi_dest":  "453.1_SERIE_DESEADA_0_0_24_58",
    "isac":      "33.2_ISAC_SIN_EDAD_0_M_23_56",
    "ica_expo":  "74.3_IET_0_M_16",
    "ica_pp":    "74.3_IEPP_0_M_35",
    "ica_moi":   "74.3_IEMOI_0_M_46",
    "ica_cye":   "74.3_IECE_0_M_35",
    "eph_des":   "45.2_ECTDT_0_T_33",   # Tasa desocupación total (EPH continua)
    "eph_act":   "43.2_ECTAT_0_T_33",   # Tasa actividad total
    "eph_emp":   "44.2_ECTET_0_T_30",   # Tasa empleo total
    "ripte":     "158.1_REPTE_0_0_5",
    "rec":       "172.3_TL_RECAION_M_0_0_17",
    "bm":        "90.1_BMT_0_0_20",
    # === nuevos ===
    "sal_priv_reg":  "149.1_SOR_PRIADO_OCTU_0_25",   # Índice salarios privado registrado
    "sal_priv_no":   "149.1_SOR_PRIADO_OCTU_0_28",   # Índice salarios privado no registrado
    "sal_pub":       "149.1_SOR_PUBICO_OCTU_0_14",   # Índice salarios público
    "empleo_sipa":   "151.1_TL_ESTADAD_2012_M_20",   # Total trabajadores SIPA (con est) — en miles
    "empleo_priv":   "151.1_AARIADODAD_2012_M_31",   # Asalariados sector privado (con est) — en miles
    "turismo_rec":   "322.3_TURISMO_REIVO__17",      # Turismo receptivo (Ezeiza+Aeroparque)
    "turismo_em":    "322.3_TURISMO_EMIVO__15",      # Turismo emisivo
    "sal_primario":  "452.3_RESULTADO_RIO_0_M_18_54", # IMIG resultado primario mensual SPNF
    "sal_financiero":"378.9_RESULTADO_017_0_M_18_90", # Resultado financiero (met 2017) mensual
}

# UCI sectorial (INDEC publica por sector; nivel general "tot" sale del PDF y lo dejamos del histórico)
UCI_SECTORES = {
    "text":   "31.3_UPT_2004_M_23",    # textiles
    "quim":   "31.3_USPQ_2004_M_34",   # químicos
    "metal":  "31.3_UIMB_2004_M_33",   # metales básicos
}

# Variables BCRA usadas
BCRA_VARS = {
    "reservas":      1,    # Reservas internacionales BRUTAS (USD M, diaria)
    "reservas_xdeg": 74,   # Reservas excl asignaciones DEG 2009 (USD M, diaria)
    "deg_2009":      83,   # Asignaciones de DEGS del año 2009 (USD M)
    "pase_bis":      76,   # Divisas-pase pasivo en dólares con el exterior (BIS-type)
    "tc_minor":      4,    # TC minorista vendedor, diaria
    "tc_mayor":      5,    # TC mayorista de referencia, diaria
    "bm_diaria":     15,   # Base monetaria diaria (M$)
    "ipc_vm":        27,   # Variación mensual IPC
    "ipc_via":       28,   # Variación interanual IPC
}

TIMEOUT = 30
RETRIES = 3
HEADERS = {"Accept-Language": "es-AR", "User-Agent": "boludos-ppt-scraper/1.0"}

# ──────────────────────────────────────────────────────────────────────────────
# Helpers de logging
# ──────────────────────────────────────────────────────────────────────────────
def log(msg: str, level: str = "info") -> None:
    ts = datetime.now(timezone.utc).strftime("%H:%M:%S")
    icon = {"info": "  ", "ok": "OK", "warn": "!!", "err": "XX"}.get(level, "  ")
    print(f"[{ts}] {icon} {msg}", flush=True)

# ──────────────────────────────────────────────────────────────────────────────
# HTTP helpers
# ──────────────────────────────────────────────────────────────────────────────
def get_json(url: str, params: dict | None = None, headers: dict | None = None) -> Any:
    """GET con retries y backoff exponencial."""
    last_err = None
    for attempt in range(RETRIES):
        try:
            r = requests.get(url, params=params, headers=headers or HEADERS, timeout=TIMEOUT)
            r.raise_for_status()
            return r.json()
        except (requests.RequestException, ValueError) as e:
            last_err = e
            if attempt < RETRIES - 1:
                time.sleep(2 ** attempt)
    raise RuntimeError(f"GET {url} fallido tras {RETRIES} intentos: {last_err}")

# ──────────────────────────────────────────────────────────────────────────────
# Fetchers por fuente
# ──────────────────────────────────────────────────────────────────────────────
def fetch_indec(series_ids: list[str] | str, last: int = 60) -> list[list]:
    ids = ",".join(series_ids) if isinstance(series_ids, list) else series_ids
    data = get_json(INDEC_API, params={"ids": ids, "last": last, "format": "json"})
    rows = data.get("data", [])
    # API ya devuelve ASC (oldest first). Forzamos sort por las dudas.
    return sorted(rows, key=lambda r: r[0] or "")

def fetch_bcra(var_id: int, limit: int = 365) -> list[dict]:
    """Devuelve lista de {fecha, valor} ordenada cronológicamente."""
    data = get_json(f"{BCRA_API}{var_id}", params={"limit": limit})
    detalle = data.get("results", [{}])[0].get("detalle", [])
    return list(reversed(detalle))

def fetch_embi_historico() -> list[dict]:
    """Devuelve serie completa EMBI desde 1999."""
    return get_json(f"{ARG_DATOS}/finanzas/indices/riesgo-pais")

def fetch_embi_ultimo() -> dict | None:
    """La API a veces omite 'valor' en los registros recientes; /ultimo siempre lo trae."""
    try:
        return get_json(f"{ARG_DATOS}/finanzas/indices/riesgo-pais/ultimo")
    except Exception as e:
        log(f"EMBI /ultimo falló: {e}", "warn")
        return None

def fetch_cotizaciones() -> dict[str, list[dict]]:
    """Blue, MEP, CCL históricos de ArgentinaDatos."""
    out = {}
    for tipo in ("blue", "bolsa", "contadoconliqui"):
        try:
            data = get_json(f"{ARG_DATOS}/cotizaciones/dolares/{tipo}")
            out[tipo] = data
        except Exception as e:
            log(f"Cotización {tipo} falló: {e}", "warn")
            out[tipo] = []
    return out

def fetch_bluelytics() -> dict | None:
    try:
        return get_json(BLUELYTICS)
    except Exception as e:
        log(f"Bluelytics falló: {e}", "warn")
        return None

# ──────────────────────────────────────────────────────────────────────────────
# Upsert helpers
# ──────────────────────────────────────────────────────────────────────────────
def upsert_by_f(arr: list[dict], record: dict) -> bool:
    """Inserta o actualiza por clave 'f'. Devuelve True si fue insert."""
    f = record["f"]
    for i, r in enumerate(arr):
        if r.get("f") == f:
            # actualizar solo campos nuevos / cambiados; preservar lo que ya estaba
            for k, v in record.items():
                if v is not None:
                    r[k] = v
            r.pop("proj", None)  # si lo estamos actualizando con datos reales, ya no es proyección
            return False
    arr.append(record)
    arr.sort(key=lambda x: x.get("f", ""))
    return True

def ym(fecha: str) -> str:
    """YYYY-MM-DD → YYYY-MM"""
    return fecha[:7]

# ──────────────────────────────────────────────────────────────────────────────
# Procesadores por serie
# ──────────────────────────────────────────────────────────────────────────────
def update_ipc(D: dict) -> int:
    """IPC: usa INDEC para nivel n; recalcula vm y via."""
    arr = D.setdefault("ipc", [])
    nuevos = 0

    # 1) Traer nivel desde INDEC (datos oficiales con desglose si vienen)
    indec_last = None
    try:
        rows = fetch_indec(INDEC_SERIES["ipc"], last=24)
        for row in rows:
            if not row[0] or row[1] is None:
                continue
            f = ym(row[0])
            n = round(row[1], 4)
            rec = next((x for x in arr if x.get("f") == f), None)
            if rec is None:
                arr.append({"f": f, "n": n})
                nuevos += 1
            else:
                # Si era proyección, ahora es real: limpiamos proj y reemplazamos n
                rec["n"] = n
                rec.pop("proj", None)
        if rows:
            indec_last = rows[-1][0][:7]
            log(f"IPC INDEC: {nuevos} períodos nuevos (último {indec_last})", "ok")
    except Exception as e:
        log(f"IPC INDEC API falló: {e}", "warn")

    # 2) Completar meses recientes que INDEC aún no publica vía BCRA Var 27 (vm) y 28 (via)
    #    La BCRA suele tener el dato 1 mes antes que el API datos.gob.ar
    try:
        vm_rows = fetch_bcra(BCRA_VARS["ipc_vm"], limit=24)
        via_rows = fetch_bcra(BCRA_VARS["ipc_via"], limit=24)
        via_by_month = {ym(v["fecha"]): float(v["valor"]) for v in via_rows if v.get("fecha")}
        # solo procesar meses POSTERIORES al último de INDEC, o que estén marcados como proj
        for vm_row in vm_rows[-6:]:
            fecha = vm_row.get("fecha", "")
            if not fecha:
                continue
            f = ym(fecha)
            vm = float(vm_row["valor"])
            via = via_by_month.get(f)
            rec = next((x for x in arr if x.get("f") == f), None)
            es_posterior_indec = indec_last is None or f > indec_last
            es_proj = rec is not None and rec.get("proj")
            if rec is None or es_proj:
                if not es_posterior_indec and not es_proj:
                    continue  # INDEC ya tiene este mes con datos reales, no tocamos
                # Necesitamos calcular n desde el mes anterior REAL (no proj)
                prev_real = None
                for x in reversed(arr):
                    if x.get("f", "") < f and x.get("n") and not x.get("proj"):
                        prev_real = x
                        break
                if prev_real:
                    n_calc = round(prev_real["n"] * (1 + vm / 100), 4)
                    if rec is None:
                        arr.append({"f": f, "n": n_calc, "vm": round(vm, 2),
                                    "via": round(via, 2) if via is not None else None})
                        nuevos += 1
                    else:
                        # reemplazar proyección con dato real BCRA
                        rec.clear()
                        rec.update({"f": f, "n": n_calc, "vm": round(vm, 2)})
                        if via is not None:
                            rec["via"] = round(via, 2)
                    log(f"IPC {f}: completado vía BCRA ({vm:.2f}% mensual, n={n_calc:.2f})", "ok")
    except Exception as e:
        log(f"IPC BCRA Var27/28 falló: {e}", "warn")

    # 3) Recalcular vm/via para todos los registros con n
    arr.sort(key=lambda x: x.get("f", ""))
    idx = {x["f"]: i for i, x in enumerate(arr) if "f" in x}
    for r in arr:
        if r.get("n") is None or r.get("proj"):
            continue
        f = r["f"]
        prev_f = prev_month(f)
        py_f = prev_year(f)
        if prev_f in idx:
            p = arr[idx[prev_f]]
            if p.get("n") and not p.get("proj"):
                r["vm"] = round((r["n"] / p["n"] - 1) * 100, 2)
        if py_f in idx:
            p12 = arr[idx[py_f]]
            if p12.get("n") and not p12.get("proj"):
                r["via"] = round((r["n"] / p12["n"] - 1) * 100, 2)

    return nuevos

# Divisiones COICOP del IPC → labels cortos para charts
IPC_DIV_LBL = {
    "01": "Alimentos y bebidas",
    "02": "Bebidas alc. y tabaco",
    "03": "Prendas de vestir",
    "04": "Vivienda y servicios",
    "05": "Equip. del hogar",
    "06": "Salud",
    "07": "Transporte",
    "08": "Comunicación",
    "09": "Recreación y cultura",
    "10": "Educación",
    "11": "Restaurantes y hoteles",
    "12": "Bienes y serv. varios",
}
# Mapeo división COICOP → campos legacy del schema D.ipc (drill por rubro del HTML)
IPC_DIV_LEGACY = {"01": "alim_vm", "03": "vest", "04": "viv", "06": "salud",
                  "07": "transp", "10": "educ", "11": "rest"}

def update_ipc_divisiones(D: dict) -> int:
    """IPC por división COICOP desde el CSV oficial INDEC (cobertura Nacional).

    Fuente: serie_ipc_divisiones.csv — se actualiza el mismo día que el IPC.
    Guarda en cada registro D.ipc[f].div = {"01": {vm, via, n}, ...} desde 2025-01
    (suficiente para acumulados del año en curso) y rellena los campos legacy
    (alim_vm, vest, viv, ...) que la API ya no trae.
    """
    arr = D.setdefault("ipc", [])
    url = f"{INDEC_XLS_BASE}/serie_ipc_divisiones.csv"
    try:
        r = http_get(url, serie="ipc")
        text = r.content.decode("utf-8", errors="replace")
    except Exception as e:
        log(f"IPC divisiones CSV falló: {e}", "warn")
        return 0

    def _f(v):
        v = v.strip().replace(",", ".")
        if not v or v.upper() == "NA":
            return None
        try:
            return float(v)
        except ValueError:
            return None

    por_mes: dict[str, dict] = {}
    for raw in text.strip().split("\n")[1:]:
        p = [x.strip() for x in raw.split(";")]
        if len(p) < 8:
            continue
        cod, desc, clasif, per, idx, vm, via, reg = p[:8]
        if reg != "Nacional" or "COICOP" not in clasif or cod not in IPC_DIV_LBL:
            continue
        if len(per) != 6 or not per.isdigit():
            continue
        f = f"{per[:4]}-{per[4:6]}"
        if f < "2025-01":
            continue
        d = {}
        n_v, vm_v, via_v = _f(idx), _f(vm), _f(via)
        if n_v is not None:   d["n"] = round(n_v, 2)
        if vm_v is not None:  d["vm"] = round(vm_v, 2)
        if via_v is not None: d["via"] = round(via_v, 2)
        if d:
            por_mes.setdefault(f, {})[cod] = d

    actualizados = 0
    for f, divs in por_mes.items():
        rec = next((x for x in arr if x.get("f") == f), None)
        if rec is None or rec.get("proj"):
            continue
        rec["div"] = divs
        # Rellenar campos legacy (drill por rubro) si faltan
        for cod, campo in IPC_DIV_LEGACY.items():
            vm_v = divs.get(cod, {}).get("vm")
            if vm_v is not None and rec.get(campo) is None:
                rec[campo] = vm_v
        via_alim = divs.get("01", {}).get("via")
        if via_alim is not None and rec.get("alim_via") is None:
            rec["alim_via"] = via_alim
        actualizados += 1
    if actualizados:
        ult = max(por_mes.keys())
        log(f"IPC divisiones (CSV oficial): {actualizados} meses con desglose COICOP (último {ult})", "ok")
    return actualizados

def update_emae(D: dict) -> int:
    arr = D.setdefault("emae", [])
    nuevos = 0
    try:
        rows = fetch_indec([INDEC_SERIES["emae_orig"], INDEC_SERIES["emae_dest"]], last=36)
        for row in rows:
            if not row[0]:
                continue
            f = ym(row[0])
            orig = row[1]
            dest = row[2]
            if orig is None:
                continue
            rec = next((x for x in arr if x.get("f") == f), None)
            new = {
                "f": f,
                "orig": round(orig, 2),
                "dest": round(dest, 2) if dest is not None else None,
            }
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                rec.update({k: v for k, v in new.items() if v is not None})
        # recalcular via (interanual sobre orig) y vm (mensual sobre dest)
        arr.sort(key=lambda x: x.get("f", ""))
        idx = {x["f"]: i for i, x in enumerate(arr)}
        for i, r in enumerate(arr):
            py = prev_year(r["f"])
            if py in idx and arr[idx[py]].get("orig"):
                r["via"] = round((r["orig"] / arr[idx[py]]["orig"] - 1) * 100, 2)
            if i > 0 and r.get("dest") and arr[i - 1].get("dest"):
                r["vm"] = round((r["dest"] / arr[i - 1]["dest"] - 1) * 100, 2)
        log(f"EMAE: {nuevos} períodos nuevos (último {rows[-1][0][:7]})", "ok")
    except Exception as e:
        log(f"EMAE falló: {e}", "warn")
    return nuevos

def update_ipi_isac(D: dict) -> int:
    nuevos = 0
    # IPI
    try:
        rows = fetch_indec([INDEC_SERIES["ipi_orig"], INDEC_SERIES["ipi_dest"]], last=36)
        arr = D.setdefault("ipi", [])
        for row in rows:
            if not row[0] or row[1] is None:
                continue
            f = ym(row[0])
            rec = next((x for x in arr if x.get("f") == f), None)
            new = {"f": f, "orig": round(row[1], 2)}
            if row[2] is not None:
                new["dest"] = round(row[2], 2)
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                rec.update({k: v for k, v in new.items() if v is not None})
        # recalc via/vm
        recalc_via_vm(arr, key_orig="orig", key_dest="dest")
        log(f"IPI: actualizado (último {rows[-1][0][:7]})", "ok")
    except Exception as e:
        log(f"IPI falló: {e}", "warn")
    # ISAC
    try:
        rows = fetch_indec(INDEC_SERIES["isac"], last=36)
        arr = D.setdefault("isac", [])
        for row in rows:
            if not row[0] or row[1] is None:
                continue
            f = ym(row[0])
            rec = next((x for x in arr if x.get("f") == f), None)
            new = {"f": f, "dest": round(row[1], 2)}
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                rec.update(new)
        arr.sort(key=lambda x: x["f"])
        for i, r in enumerate(arr):
            if i > 0 and r.get("dest") and arr[i - 1].get("dest"):
                r["vm"] = round((r["dest"] / arr[i - 1]["dest"] - 1) * 100, 2)
        log(f"ISAC: actualizado (último {rows[-1][0][:7]})", "ok")
    except Exception as e:
        log(f"ISAC falló: {e}", "warn")
    return nuevos

def update_bcra_monthly(D: dict) -> int:
    """Reservas, TC, BM — todas mensuales, último día observado del mes."""
    nuevos = 0

    # Reservas (mensual): BRUTAS + excl DEG + pase BIS + DEG → calcula netas aproximadas
    try:
        # Fetch en paralelo para velocidad
        rows_brutas = fetch_bcra(BCRA_VARS["reservas"], limit=730)        # 2 años diarios
        rows_xdeg   = fetch_bcra(BCRA_VARS["reservas_xdeg"], limit=730)
        rows_pase   = fetch_bcra(BCRA_VARS["pase_bis"], limit=730)
        rows_deg    = fetch_bcra(BCRA_VARS["deg_2009"], limit=730)

        def by_month_last(rows):
            """Agrupa por mes (YYYY-MM), retorna dict {f: valor del último día observado}."""
            out = {}
            for r in rows:
                f = ym(r["fecha"])
                try:
                    out[f] = float(r["valor"])
                except (TypeError, ValueError):
                    pass
            return out

        m_brutas = by_month_last(rows_brutas)
        m_xdeg   = by_month_last(rows_xdeg)
        m_pase   = by_month_last(rows_pase)
        m_deg    = by_month_last(rows_deg)

        arr = D.setdefault("reservas", [])
        all_meses = set(m_brutas) | set(m_xdeg) | set(m_pase) | set(m_deg)
        for f in all_meses:
            brutas = m_brutas.get(f)
            xdeg = m_xdeg.get(f)
            pase = m_pase.get(f, 0.0)
            deg = m_deg.get(f, 0.0)
            # netas_aprox = brutas - pase BIS - DEG 2009
            # NOTA: no incluye encajes USD ni swap China — para eso hace falta el balance semanal BCRA
            netas_aprox = None
            if brutas is not None:
                netas_aprox = brutas - (pase or 0) - (deg or 0)

            rec = next((x for x in arr if x.get("f") == f), None)
            new = {"f": f}
            if brutas is not None:
                new["v"] = round(brutas)        # Brutas (compatibilidad con schema actual)
                new["brutas"] = round(brutas)
            if xdeg is not None:
                new["excl_deg"] = round(xdeg)
            if pase:
                new["pase_bis"] = round(pase)
            if deg:
                new["deg"] = round(deg)
            if netas_aprox is not None:
                new["netas_aprox"] = round(netas_aprox)
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                rec.update({k: v for k, v in new.items() if k != "f"})
        arr.sort(key=lambda x: x["f"])
        if rows_brutas:
            log(f"Reservas BCRA: último {rows_brutas[-1]['fecha']} = USD {rows_brutas[-1]['valor']:.0f}M (brutas) + componentes", "ok")
    except Exception as e:
        log(f"Reservas BCRA falló: {e}", "warn")

    # TC Oficial (mensual)
    try:
        rows = fetch_bcra(BCRA_VARS["tc_minor"], limit=365)
        arr = D.setdefault("tc", [])
        by_month = {}
        for r in rows:
            by_month[ym(r["fecha"])] = r
        for f, r in by_month.items():
            v = round(float(r["valor"]), 2)
            rec = next((x for x in arr if x.get("f") == f), None)
            if rec is None:
                arr.append({"f": f, "of": v})
            else:
                rec["of"] = v
        arr.sort(key=lambda x: x["f"])
        # Recalcular variaciones
        for i, r in enumerate(arr):
            if i > 0 and r.get("of") and arr[i - 1].get("of"):
                r["vm_of"] = round((r["of"] / arr[i - 1]["of"] - 1) * 100, 2)
            py = prev_year(r["f"])
            py_rec = next((x for x in arr if x.get("f") == py), None)
            if py_rec and py_rec.get("of"):
                r["via_of"] = round((r["of"] / py_rec["of"] - 1) * 100, 2)
        log(f"TC Oficial BCRA: último {rows[-1]['fecha']} = ${rows[-1]['valor']:.2f}", "ok")
    except Exception as e:
        log(f"TC Oficial BCRA falló: {e}", "warn")

    # Base monetaria (mensual)
    try:
        rows = fetch_bcra(BCRA_VARS["bm_diaria"], limit=365)
        arr = D.setdefault("bm", [])
        by_month = {}
        for r in rows:
            by_month[ym(r["fecha"])] = r
        for f, r in by_month.items():
            v = round(float(r["valor"]))
            rec = next((x for x in arr if x.get("f") == f), None)
            if rec is None:
                arr.append({"f": f, "tot": v})
            else:
                rec["tot"] = v
        arr.sort(key=lambda x: x["f"])
        log(f"Base monetaria BCRA: último {rows[-1]['fecha']}", "ok")
    except Exception as e:
        log(f"BM BCRA falló: {e}", "warn")

    return nuevos

def update_embi(D: dict) -> int:
    arr = D.setdefault("embi", [])
    nuevos = 0
    try:
        rp = fetch_embi_historico()
        by_month = {}
        for x in rp:
            if not x.get("fecha"):
                continue
            v = x.get("valor")
            if v is None:
                continue
            by_month[ym(x["fecha"])] = v
        for f, v in by_month.items():
            rec = next((x for x in arr if x.get("f") == f), None)
            if rec is None:
                arr.append({"f": f, "v": v})
                nuevos += 1
            else:
                rec["v"] = v
        arr.sort(key=lambda x: x["f"])
        # Completar último con /ultimo (la API a veces omite valor en los más recientes)
        ult = fetch_embi_ultimo()
        if ult and ult.get("valor"):
            f = ym(ult["fecha"])
            rec = next((x for x in arr if x.get("f") == f), None)
            if rec is None:
                arr.append({"f": f, "v": ult["valor"]})
                nuevos += 1
            else:
                rec["v"] = ult["valor"]
            arr.sort(key=lambda x: x["f"])
            log(f"EMBI: último {ult['fecha']} = {ult['valor']} bps", "ok")
    except Exception as e:
        log(f"EMBI falló: {e}", "warn")
    return nuevos

def update_simple_monthly(D: dict, key: str, series_id: str, val_key: str, label: str) -> int:
    """Helper para series mensuales simples con un solo valor."""
    arr = D.setdefault(key, [])
    nuevos = 0
    try:
        rows = fetch_indec(series_id, last=36)
        for row in rows:
            if not row[0] or row[1] is None:
                continue
            f = ym(row[0])
            v = round(float(row[1]), 2)
            rec = next((x for x in arr if x.get("f") == f), None)
            if rec is None:
                arr.append({"f": f, val_key: v})
                nuevos += 1
            else:
                rec[val_key] = v
        arr.sort(key=lambda x: x["f"])
        log(f"{label}: actualizado (último {rows[-1][0][:7]})", "ok")
    except Exception as e:
        log(f"{label} falló: {e}", "warn")
    return nuevos

# ──────────────────────────────────────────────────────────────────────────────
# Utilidades de fecha
# ──────────────────────────────────────────────────────────────────────────────
def prev_month(f: str) -> str:
    y, m = int(f[:4]), int(f[5:7])
    if m == 1:
        return f"{y - 1:04d}-12"
    return f"{y:04d}-{m - 1:02d}"

def prev_year(f: str) -> str:
    return f"{int(f[:4]) - 1:04d}-{f[5:7]}"

def recalc_via_vm(arr: list[dict], key_orig: str = "orig", key_dest: str = "dest") -> None:
    arr.sort(key=lambda x: x.get("f", ""))
    idx = {x["f"]: i for i, x in enumerate(arr) if "f" in x}
    for r in arr:
        py = prev_year(r["f"])
        if py in idx and arr[idx[py]].get(key_orig):
            try:
                r["via"] = round((r[key_orig] / arr[idx[py]][key_orig] - 1) * 100, 2)
            except (KeyError, TypeError):
                pass
        i = idx[r["f"]]
        if i > 0 and r.get(key_dest) and arr[i - 1].get(key_dest):
            try:
                r["vm"] = round((r[key_dest] / arr[i - 1][key_dest] - 1) * 100, 2)
            except (KeyError, TypeError):
                pass

# ──────────────────────────────────────────────────────────────────────────────
# Procesadores adicionales: UCI, Salarios, Empleo, Turismo, MERVAL
# ──────────────────────────────────────────────────────────────────────────────
def update_uci_sectorial(D: dict) -> int:
    """UCI por sector (textiles, químicos, metales básicos). Nivel general 'tot' del histórico."""
    arr = D.setdefault("uci", [])
    nuevos = 0
    # Mapeo nuestro key → series INDEC
    rows_por_key = {}
    for key, series_id in UCI_SECTORES.items():
        try:
            data = fetch_indec(series_id, last=36)
            for row in data:
                if not row[0] or row[1] is None:
                    continue
                f = ym(row[0])
                rows_por_key.setdefault(f, {})[key] = round(float(row[1]), 1)
        except Exception as e:
            log(f"UCI {key} falló: {e}", "warn")

    for f, vals in rows_por_key.items():
        rec = next((x for x in arr if x.get("f") == f), None)
        if rec is None:
            new = {"f": f, **vals}
            arr.append(new)
            nuevos += 1
        else:
            rec.update(vals)
    arr.sort(key=lambda x: x["f"])
    if rows_por_key:
        ult = max(rows_por_key.keys())
        log(f"UCI sectorial: actualizado (último {ult})", "ok")
    return nuevos

def update_salarios(D: dict) -> int:
    """Salarios privado registrado / no registrado / público. Schema: {f, is_r, real_pub, real_priv}."""
    arr = D.setdefault("salarios", [])
    nuevos = 0
    try:
        ids = [INDEC_SERIES["sal_priv_reg"], INDEC_SERIES["sal_priv_no"], INDEC_SERIES["sal_pub"]]
        data = fetch_indec(ids, last=24)
        for row in data:
            if not row[0]:
                continue
            f = ym(row[0])
            is_r = row[1]      # privado registrado (lo usamos como índice principal)
            priv_no = row[2]
            pub = row[3]
            if is_r is None and priv_no is None and pub is None:
                continue
            rec = next((x for x in arr if x.get("f") == f), None)
            new = {"f": f}
            if is_r is not None:
                new["is_r"] = round(float(is_r), 2)
            if priv_no is not None:
                new["real_priv"] = round(float(priv_no), 2)
            if pub is not None:
                new["real_pub"] = round(float(pub), 2)
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                rec.update(new)
        arr.sort(key=lambda x: x["f"])
        if data:
            log(f"Salarios: actualizado (último {data[-1][0][:7]})", "ok")
    except Exception as e:
        log(f"Salarios falló: {e}", "warn")
    return nuevos

def update_comercio_interior(D: dict) -> int:
    """Comercio interior — Supermercados + Autoservicios Mayoristas (INDEC).

    Fuentes:
      - Supermercados:        serie_supermercados.xlsx Cuadro 2 (precios constantes)
      - Autoservicios mayor.: sh_super_mayoristas.xls Cuadro 8 (precios constantes)

    Schema:
      D.super  = [{f, n_real, via_real, acum_real, dest_real, vm_dest}]
      D.mayor  = [{f, n_real, via_real, acum_real, dest_real, vm_dest}]
    """
    nuevos = 0
    # === Supermercados (XLSX — Cuadro 1: precios CONSTANTES, no corrientes) ===
    try:
        import openpyxl
        url = "https://www.indec.gob.ar/ftp/cuadros/economia/serie_supermercados.xlsx"
        r = http_get(url, serie="super")
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
        # Cuadro 1 = Índice ventas totales a PRECIOS CONSTANTES (real, base 2017=100)
        # Cuadro 2 era corrientes (nominal) y daba lecturas equivocadas
        sn = next((s for s in wb.sheetnames if "Cuadro 1" in s), None)
        if sn:
            ws = wb[sn]
            arr = D.setdefault("super", [])
            for row_idx in range(7, ws.max_row + 1):
                periodo = ws.cell(row_idx, 1).value
                if periodo is None:
                    continue
                if hasattr(periodo, "strftime"):
                    f = periodo.strftime("%Y-%m")
                elif isinstance(periodo, str) and len(periodo) >= 7 and "-" in periodo:
                    f = periodo[:7]
                else:
                    continue
                n_real    = _safe_float(ws.cell(row_idx, 2).value)
                via_real  = _safe_float(ws.cell(row_idx, 3).value)
                acum_real = _safe_float(ws.cell(row_idx, 4).value)
                dest_real = _safe_float(ws.cell(row_idx, 6).value)
                vm_dest   = _safe_float(ws.cell(row_idx, 7).value)
                if n_real is None:
                    continue
                rec = next((x for x in arr if x.get("f") == f), None)
                new = {"f": f, "n_real": round(n_real, 2)}
                if via_real is not None:  new["via_real"]  = round(via_real, 2)
                if acum_real is not None: new["acum_real"] = round(acum_real, 2)
                if dest_real is not None: new["dest_real"] = round(dest_real, 2)
                if vm_dest is not None:   new["vm_dest"]   = round(vm_dest, 2)
                if rec is None:
                    arr.append(new)
                    nuevos += 1
                else:
                    rec.update({k: v for k, v in new.items() if k != "f"})
            arr.sort(key=lambda x: x["f"])
            if arr:
                log(f"Supermercados (XLSX): actualizado (último {arr[-1]['f']}, vía real={arr[-1].get('via_real','—')}%)", "ok")
    except Exception as e:
        log(f"Supermercados XLSX falló: {e}", "warn")

    # === Autoservicios Mayoristas (XLS — Cuadro 9 ventas totales) ===
    # Estructura: col 0=tag, col 1=año, col 2=mes, col 3=ventas_nom (miles $),
    #             col 4=bocas, col 5=ventas/boca, col 6=superficie, col 7=ventas/m2,
    #             col 8=operaciones, col 9=ventas/operación
    try:
        url = "https://www.indec.gob.ar/ftp/cuadros/economia/sh_super_mayoristas.xls"
        r = http_get(url, serie="mayor")
        wb = xlrd.open_workbook(file_contents=r.content)
        try:
            sh = wb.sheet_by_name("Cuadro 9")
        except Exception:
            sh = None
        if sh is not None:
            arr = D.setdefault("mayor", [])
            current_year = None
            ult_f = None
            for r2 in range(sh.nrows):
                y_cell = sh.cell_value(r2, 1)
                m_cell = sh.cell_value(r2, 2)
                if isinstance(y_cell, float) and y_cell > 1990:
                    current_year = int(y_cell)
                elif isinstance(y_cell, str):
                    s = y_cell.strip().rstrip("*").strip()
                    if s.isdigit() and 1990 < int(s) < 2100:
                        current_year = int(s)
                if not current_year:
                    continue
                month_num = _month_num(m_cell) if m_cell else None
                if not month_num:
                    continue
                ventas_nom = _safe_float(sh.cell_value(r2, 3))
                bocas      = _safe_float(sh.cell_value(r2, 4)) if sh.ncols > 4 else None
                vta_boca   = _safe_float(sh.cell_value(r2, 5)) if sh.ncols > 5 else None
                operac     = _safe_float(sh.cell_value(r2, 8)) if sh.ncols > 8 else None
                if ventas_nom is None:
                    continue
                f = f"{current_year:04d}-{month_num:02d}"
                ult_f = f
                rec = next((x for x in arr if x.get("f") == f), None)
                new = {"f": f, "ventas_nom": int(round(ventas_nom))}
                if bocas is not None:    new["bocas"]    = int(round(bocas))
                if vta_boca is not None: new["vta_boca"] = round(vta_boca, 0)
                if operac is not None:   new["operac"]   = int(round(operac))
                if rec is None:
                    arr.append(new)
                    nuevos += 1
                else:
                    rec.update({k: v for k, v in new.items() if k != "f"})
            arr.sort(key=lambda x: x["f"])
            # Calcular variación interanual nominal contra mismo mes año anterior
            idx = {x["f"]: i for i, x in enumerate(arr) if "f" in x}
            for rec in arr:
                py = prev_year(rec["f"])
                if py in idx:
                    prev = arr[idx[py]]
                    if prev.get("ventas_nom") and rec.get("ventas_nom"):
                        rec["via_nom"] = round((rec["ventas_nom"] / prev["ventas_nom"] - 1) * 100, 2)
            if ult_f:
                log(f"Autoservicios mayoristas (Cuadro 9): actualizado (último {ult_f})", "ok")
    except Exception as e:
        log(f"Mayoristas XLS falló: {e}", "warn")
    return nuevos

def update_salarios_csv(D: dict) -> int:
    """Índice de Salarios desde el CSV oficial INDEC (más fresco que la API series).

    Fuente: indec.gob.ar/ftp/cuadros/sociedad/indice_salarios.csv
    Columnas:
      - periodo (DD/MM/YYYY)
      - IS_sector_privado_registrado
      - IS_sector_publico
      - IS_total_registrado
      - IS_sector_no_registrado
      - IS_indice_total

    Schema D.salarios (mantener compatibilidad):
      is_r       ← IS_total_registrado
      real_priv  ← IS_sector_privado_registrado (nominal, se deflacta en HTML)
      real_pub   ← IS_sector_publico (nominal)
      no_reg     ← IS_sector_no_registrado (nuevo)
      ind_tot    ← IS_indice_total (nuevo)
    """
    arr = D.setdefault("salarios", [])
    nuevos = 0
    url = "https://www.indec.gob.ar/ftp/cuadros/sociedad/indice_salarios.csv"
    try:
        r = http_get(url, serie="salarios")
        text = r.content.decode("utf-8", errors="replace")
        lines = text.strip().split("\n")
        # Parsear header
        header = [h.strip() for h in lines[0].split(";")]
        # Mapear nombres de columnas a claves del schema
        COL_MAP = {
            "IS_sector_privado_registrado": "real_priv",
            "IS_sector_publico":            "real_pub",
            "IS_total_registrado":          "is_r",
            "IS_sector_no_registrado":      "no_reg",
            "IS_indice_total":              "ind_tot",
        }
        col_idx = {COL_MAP[h]: i for i, h in enumerate(header) if h in COL_MAP}
        if "is_r" not in col_idx:
            log("Salarios CSV: header inesperado", "warn")
            return 0
        for line in lines[1:]:
            parts = [p.strip() for p in line.strip().split(";")]
            if len(parts) < 6:
                continue
            # Parsear fecha DD/MM/YYYY → YYYY-MM
            try:
                d, m, y = parts[0].split("/")
                f = f"{int(y):04d}-{int(m):02d}"
            except Exception:
                continue
            new = {"f": f}
            for key, idx in col_idx.items():
                raw = parts[idx]
                if raw in ("NA", "", "N/A"):
                    continue
                try:
                    val = float(raw.replace(",", "."))
                    new[key] = round(val, 2)
                except ValueError:
                    pass
            rec = next((x for x in arr if x.get("f") == f), None)
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                for k, v in new.items():
                    if k != "f":
                        rec[k] = v
        arr.sort(key=lambda x: x["f"])
        if arr:
            log(f"Salarios (CSV oficial): actualizado (último {arr[-1]['f']}, is_r={arr[-1].get('is_r','—')})", "ok")
    except Exception as e:
        log(f"Salarios CSV falló: {e}", "warn")
    return nuevos

def update_eph(D: dict) -> int:
    """EPH mercado de trabajo (trimestral): tasas de actividad, empleo y desocupación.

    Schema: D.eph = [{p: '1er t. 2026', td, act, emp}] (tasas en %).
    Fuente: API datos.gob.ar (EPH continua, total 31 aglomerados). La API trae
    fracciones (0.078 = 7.8%) → ×100. OJO: la API tiene rezago de semanas para
    el trimestre recién publicado; EPH_OVERRIDE inyecta el último dato oficial
    (de la web INDEC) hasta que la API lo incorpore.
    """
    QLBL = {1: "1er t.", 4: "2do t.", 7: "3er t.", 10: "4to t."}
    arr = D.setdefault("eph", [])
    nuevos = 0
    try:
        ids = [INDEC_SERIES["eph_des"], INDEC_SERIES["eph_act"], INDEC_SERIES["eph_emp"]]
        data = fetch_indec(ids, last=24)
        for row in data:
            if not row[0]:
                continue
            y, m = int(row[0][:4]), int(row[0][5:7])
            if m not in QLBL:
                continue
            p = f"{QLBL[m]} {y}"
            td = round(float(row[1]) * 100, 1) if row[1] is not None else None
            act = round(float(row[2]) * 100, 1) if row[2] is not None else None
            emp = round(float(row[3]) * 100, 1) if row[3] is not None else None
            rec = next((x for x in arr if x.get("p") == p), None)
            new = {"p": p}
            if td is not None:  new["td"] = td
            if act is not None: new["act"] = act
            if emp is not None: new["emp"] = emp
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                rec.update({k: v for k, v in new.items() if k != "p"})
        # Override del último trimestre oficial (web INDEC) si la API aún no lo trae
        for ov in EPH_OVERRIDE:
            if not any(x.get("p") == ov["p"] for x in arr):
                arr.append(dict(ov))
                nuevos += 1
        # Orden cronológico por (año, trimestre)
        qord = {"1er t.": 1, "2do t.": 2, "3er t.": 3, "4to t.": 4}
        def _key(r):
            parts = r["p"].rsplit(" ", 1)
            return (int(parts[1]), qord.get(parts[0], 0))
        arr.sort(key=_key)
        if arr:
            log(f"EPH: actualizado (último {arr[-1]['p']}, td={arr[-1].get('td')}%)", "ok")
    except Exception as e:
        log(f"EPH falló: {e}", "warn")
    return nuevos

# Último trimestre EPH oficial (web INDEC) hasta que la API datos.gob.ar lo incorpore.
# Cuando la API lo traiga, sobrescribe estos valores automáticamente.
EPH_OVERRIDE = [
    {"p": "1er t. 2026", "td": 7.8, "act": 48.6, "emp": 44.8},
]

def update_empleo(D: dict) -> int:
    """Empleo SIPA total + privado. Schema histórico: {f, tot, priv} (valores en unidades, no miles).

    IMPORTANTE: INDEC devuelve estas series en MILES de trabajadores. El bedrock está en
    UNIDADES (ej. tot=10982740 = 10.9M). Multiplicamos por 1000 para mantener consistencia.
    """
    arr = D.setdefault("trabajo", [])
    nuevos = 0
    try:
        ids = [INDEC_SERIES["empleo_sipa"], INDEC_SERIES["empleo_priv"]]
        data = fetch_indec(ids, last=48)
        for row in data:
            if not row[0]:
                continue
            f = ym(row[0])
            tot_miles = row[1]
            priv_miles = row[2]
            if tot_miles is None and priv_miles is None:
                continue
            rec = next((x for x in arr if x.get("f") == f), None)
            new = {"f": f}
            if tot_miles is not None:
                new["tot"] = int(round(float(tot_miles) * 1000))  # miles → unidades
            if priv_miles is not None:
                new["priv"] = int(round(float(priv_miles) * 1000))
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                rec.update({k: v for k, v in new.items() if k != "f"})
        arr.sort(key=lambda x: x["f"])
        # Limpiar registro corrupto histórico "2019-23" (era ruido del baked) y outliers
        arr[:] = [r for r in arr
                  if r.get("f", "").startswith("20")
                  and "-" in r["f"]
                  and len(r["f"]) == 7
                  and r.get("tot", 0) > 1000000]  # filtra el 2012-01 raro con tot=4.8M
        if data:
            log(f"Empleo SIPA: actualizado (último {data[-1][0][:7]}, tot+priv en unidades)", "ok")
    except Exception as e:
        log(f"Empleo SIPA falló: {e}", "warn")
    return nuevos

def update_turismo(D: dict) -> int:
    """Turismo internacional total (todas las vías). Schema: {f, rec, em, sal}.

    Fuente: INDEC publica `serie_turismo_receptivo_emisivo.xlsx` con el consolidado
    de todas las vías (no solo Ezeiza/Aeroparque como la serie de la API).
    """
    arr = D.setdefault("turismo", [])
    nuevos = 0
    url = f"{INDEC_XLS_BASE}/serie_turismo_receptivo_emisivo.xlsx"
    try:
        import openpyxl  # noqa: dependencia opcional
        r = http_get(url, serie="turismo")
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
        ws = wb["Turismo receptivo y emisivo"]
        # Layout: col 1 = Período (datetime), col 2 = receptivo, col 3 = emisivo, col 4 = saldo
        rows_parseadas = []
        for row_idx in range(5, ws.max_row + 1):
            periodo = ws.cell(row_idx, 1).value
            rec_v = ws.cell(row_idx, 2).value
            em_v = ws.cell(row_idx, 3).value
            if periodo is None or rec_v is None or em_v is None:
                continue
            # periodo puede ser datetime o string
            if hasattr(periodo, "strftime"):
                f = periodo.strftime("%Y-%m")
            else:
                f = str(periodo)[:7]
            rec_int = int(round(float(rec_v)))
            em_int = int(round(float(em_v)))
            sal_int = rec_int - em_int
            rec_existing = next((x for x in arr if x.get("f") == f), None)
            if rec_existing is None:
                arr.append({"f": f, "rec": rec_int, "em": em_int, "sal": sal_int})
                nuevos += 1
            else:
                rec_existing["rec"] = rec_int
                rec_existing["em"] = em_int
                rec_existing["sal"] = sal_int
            rows_parseadas.append(f)
        arr.sort(key=lambda x: x["f"])
        if rows_parseadas:
            log(f"Turismo (XLSX oficial): actualizado (último {rows_parseadas[-1]}, todas las vías)", "ok")
    except Exception as e:
        log(f"Turismo XLSX falló, fallback a API series Ezeiza+Aeroparque: {e}", "warn")
        # Fallback: serie de aeropuertos (datos chicos, sólo Ezeiza+Aeroparque)
        try:
            ids = [INDEC_SERIES["turismo_rec"], INDEC_SERIES["turismo_em"]]
            data = fetch_indec(ids, last=36)
            for row in data:
                if not row[0]:
                    continue
                f = ym(row[0])
                rec_v, em_v = row[1], row[2]
                if rec_v is None and em_v is None:
                    continue
                rec_existing = next((x for x in arr if x.get("f") == f), None)
                new = {"f": f}
                if rec_v is not None:
                    new["rec"] = int(round(float(rec_v)))
                if em_v is not None:
                    new["em"] = int(round(float(em_v)))
                if "rec" in new and "em" in new:
                    new["sal"] = new["rec"] - new["em"]
                if rec_existing is None:
                    arr.append(new)
                    nuevos += 1
                else:
                    rec_existing.update(new)
            arr.sort(key=lambda x: x["f"])
        except Exception as e2:
            log(f"Turismo fallback también falló: {e2}", "warn")
    return nuevos

def update_daily_series(D: dict) -> int:
    """Series diarias (DD): rrii_d, tc_d, embi_d, merval_d. Las KPIs del HTML leen de estas."""
    nuevos = 0

    # === Reservas diaria (BCRA Var 1) — últimos ~120 días ===
    try:
        rows = fetch_bcra(BCRA_VARS["reservas"], limit=180)
        arr = D.setdefault("rrii_d", [])
        existing = {r["f"]: i for i, r in enumerate(arr) if "f" in r}
        for r in rows:
            f = r["fecha"]
            v = round(float(r["valor"]))
            if f in existing:
                arr[existing[f]]["v"] = v
            else:
                arr.append({"f": f, "v": v})
                existing[f] = len(arr) - 1
        # Trim a últimos 180 dias
        arr.sort(key=lambda x: x["f"])
        D["rrii_d"] = arr[-180:]
        if rows:
            log(f"Reservas diaria: último {rows[-1]['fecha']} = USD {rows[-1]['valor']:.0f}M", "ok")
    except Exception as e:
        log(f"Reservas diaria falló: {e}", "warn")

    # === TC diaria (BCRA Var 4 = minorista) ===
    try:
        rows = fetch_bcra(BCRA_VARS["tc_minor"], limit=180)
        arr = D.setdefault("tc_d", [])
        existing = {r["f"]: i for i, r in enumerate(arr) if "f" in r}
        for r in rows:
            f = r["fecha"]
            v = round(float(r["valor"]), 2)
            if f in existing:
                arr[existing[f]]["of"] = v
            else:
                arr.append({"f": f, "of": v, "mep": None, "ccl": None, "blue": None})
                existing[f] = len(arr) - 1
        arr.sort(key=lambda x: x["f"])

        # Enriquecer con MEP / CCL / Blue desde ArgentinaDatos cotizaciones
        cotiz = fetch_cotizaciones()
        for tipo_key, dd_field in (("bolsa", "mep"), ("contadoconliqui", "ccl"), ("blue", "blue")):
            for x in cotiz.get(tipo_key, []):
                if not x.get("fecha"):
                    continue
                f = x["fecha"]
                v = x.get("venta") or x.get("compra")
                if v is None:
                    continue
                if f in existing:
                    arr[existing[f]][dd_field] = round(float(v), 2)
        # Bluelytics como respaldo del blue de hoy
        bl = fetch_bluelytics()
        if bl:
            f_today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            rec = next((x for x in arr if x.get("f") == f_today), None)
            if rec is None:
                rec = {"f": f_today, "of": None, "mep": None, "ccl": None, "blue": None}
                arr.append(rec)
            rec["blue"] = rec.get("blue") or round(float(bl["blue"]["value_sell"]), 2)
            rec["of"]   = rec.get("of")   or round(float(bl["oficial"]["value_sell"]), 2)
            arr.sort(key=lambda x: x["f"])
        D["tc_d"] = arr[-180:]
        log(f"TC diaria: último {arr[-1]['f']} | of={arr[-1].get('of')} blue={arr[-1].get('blue')}", "ok")
    except Exception as e:
        log(f"TC diaria falló: {e}", "warn")

    # === EMBI diaria (ArgentinaDatos riesgo-pais) — últimos 365 días ===
    try:
        rp = fetch_embi_historico()
        arr = D.setdefault("embi_d", [])
        existing = {r["f"]: i for i, r in enumerate(arr) if "f" in r}
        for x in rp:
            if not x.get("fecha"):
                continue
            v = x.get("valor")
            if v is None:
                continue
            f = x["fecha"][:10]
            if f in existing:
                arr[existing[f]]["v"] = v
            else:
                arr.append({"f": f, "v": v})
                existing[f] = len(arr) - 1
        ult = fetch_embi_ultimo()
        if ult and ult.get("valor"):
            f = ult["fecha"][:10]
            if f in existing:
                arr[existing[f]]["v"] = ult["valor"]
            else:
                arr.append({"f": f, "v": ult["valor"]})
        arr.sort(key=lambda x: x["f"])
        D["embi_d"] = arr[-365:]
        log(f"EMBI diaria: último {D['embi_d'][-1]['f']} = {D['embi_d'][-1]['v']} bps", "ok")
    except Exception as e:
        log(f"EMBI diaria falló: {e}", "warn")

    # === Índices bursátiles diarios (Yahoo Finance) — histórico completo desde 2002 ===
    # MERVAL + Dow Jones + S&P 500. Incremental: si ya hay histórico (vía continuidad
    # desde data.json previo), solo pedimos los últimos meses; si está vacío o corto,
    # bajamos todo desde 2002-01-01 con period1/period2.
    BOLSA_DESDE = datetime(2002, 1, 1, tzinfo=timezone.utc)
    for key, simbolo, nombre, escala in [
        ("merval_d", "%5EMERV", "MERVAL", 1000.0),   # MERVAL en miles (compat. bedrock)
        ("dji_d",    "%5EDJI",  "Dow Jones", 1.0),
        ("spx_d",    "%5EGSPC", "S&P 500", 1.0),
    ]:
        try:
            arr = D.setdefault(key, [])
            tiene_historico = len(arr) > 400 and arr[0].get("f", "9") <= "2003"
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{simbolo}"
            if tiene_historico:
                params = {"interval": "1d", "range": "3mo"}  # solo refrescar lo reciente
            else:
                params = {"interval": "1d",
                          "period1": int(BOLSA_DESDE.timestamp()),
                          "period2": int(datetime.now(timezone.utc).timestamp())}
            data = get_json(url, params=params, headers={**HEADERS, "User-Agent": "Mozilla/5.0"})
            result = data["chart"]["result"][0]
            ts = result["timestamp"]
            closes = result["indicators"]["quote"][0]["close"]
            existing = {r["f"]: i for i, r in enumerate(arr) if "f" in r}
            for t, c in zip(ts, closes):
                if c is None:
                    continue
                f = datetime.fromtimestamp(t, tz=timezone.utc).strftime("%Y-%m-%d")
                v = round(float(c) / escala, 2)
                if f in existing:
                    arr[existing[f]]["v"] = v
                else:
                    arr.append({"f": f, "v": v})
                    existing[f] = len(arr) - 1
            arr.sort(key=lambda x: x["f"])
            # Conservar desde 2002 en adelante (sin cap de 365)
            D[key] = [r for r in arr if r.get("f", "") >= "2002-01-01"]
            if D[key]:
                ult = D[key][-1]
                vista = f"{ult['v']}K" if key == "merval_d" else f"{ult['v']:,.0f}"
                log(f"{nombre} diaria: {len(D[key])} días (desde {D[key][0]['f']}), último {ult['f']} = {vista}", "ok")
        except Exception as e:
            log(f"{nombre} diaria falló: {e}", "warn")

    return nuevos

def update_empresas(D: dict) -> int:
    """Empresas privadas activas inscriptas en SIPA, anual.

    Fuente: OEDE (Observatorio de Empleo y Dinámica Empresarial - Ministerio de Trabajo)
    XLSX: nacional_serie_empresas_1.xlsx → Cuadro 1
    Schema: D.empresas = [{a, n, via}]
    """
    arr = D.setdefault("empresas", [])
    nuevos = 0
    url = "https://www.argentina.gob.ar/sites/default/files/nacional_serie_empresas_1.xlsx"
    try:
        import openpyxl
        r = requests.get(url, timeout=TIMEOUT, headers={**HEADERS, "User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
        ws = wb["C 1"]
        for row_idx in range(7, min(50, ws.max_row + 1)):  # datos desde fila 7
            anio_cell = ws.cell(row_idx, 1).value
            empresas_cell = ws.cell(row_idx, 2).value
            via_cell = ws.cell(row_idx, 3).value
            if anio_cell is None or empresas_cell is None:
                continue
            anio_str = str(anio_cell).rstrip("*").strip()
            if not anio_str.isdigit() or len(anio_str) != 4:
                continue
            anio = int(anio_str)
            n = int(round(float(empresas_cell)))
            via = round(float(via_cell) * 100, 2) if via_cell is not None else None
            rec = next((x for x in arr if x.get("a") == anio), None)
            new = {"a": anio, "n": n}
            if via is not None:
                new["via"] = via
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                rec.update(new)
        arr.sort(key=lambda x: x["a"])
        if arr:
            log(f"Empresas OEDE: actualizado (último {arr[-1]['a']} = {arr[-1]['n']:,} empresas)", "ok")
    except Exception as e:
        log(f"Empresas OEDE falló: {e}", "warn")
    return nuevos

def update_fiscal_hacienda(D: dict) -> int:
    """Resultado fiscal SPNF — XLSX directo de Hacienda (no IMIG INDEC).

    Hacienda publica el dato fiscal antes que la API INDEC IMIG. Los archivos están en:
      - argentina.gob.ar/sites/default/files/YYYY/MM/cuentas_publicasN.rar (RAR con XLSX adentro)
      - argentina.gob.ar/sites/default/files/{mes_corto}_{aa}.xlsx

    Esta función prueba URLs candidatas para los últimos 3 meses y, si encuentra un
    RAR, lo descomprime y procesa el XLSX 'Abril 26.xlsx' / 'Mayo 26.xlsx' etc.

    Schema: D.fiscal = [{f, sal_prim, sal_fin}] (compatible con update_fiscal_indec).
    """
    MESES_ES = ["enero","febrero","marzo","abril","mayo","junio",
                "julio","agosto","septiembre","octubre","noviembre","diciembre"]
    BASE = "https://www.argentina.gob.ar/sites/default/files"
    arr = D.setdefault("fiscal", [])
    nuevos = 0
    today = datetime.now(timezone.utc)

    def parse_y_mergear(content_xlsx, y, m):
        nonlocal nuevos
        parsed = _parse_fiscal_xlsx(content_xlsx)
        if not parsed or all(v is None for v in parsed.values()):
            return False
        f = f"{y:04d}-{m:02d}"
        rec_existing = next((x for x in arr if x.get("f") == f), None)
        new = {"f": f}
        for k, v in parsed.items():
            if v is not None:
                new[k] = round(v, 1)
        if rec_existing is None:
            arr.append(new)
            nuevos += 1
        else:
            rec_existing.update({k: v for k, v in new.items() if k != "f"})
        return True

    # Para cada mes (data) probar URLs candidatas
    for delta in range(0, 3):
        y, m = today.year, today.month - 1 - delta
        while m <= 0:
            m += 12
            y -= 1
        mes_es = MESES_ES[m - 1]
        pub_y, pub_m = y, m + 1
        if pub_m > 12:
            pub_m -= 12
            pub_y += 1
        # Formato corto (mes_yy.xlsx) — hojas 'Mes' y 'Acumulado' con valores limpios.
        # INDEC/Hacienda guarda el archivo bajo /files/{pub_y}/{pub_m}/ (subdir del mes
        # de PUBLICACIÓN). Probamos ese path primero y el legacy sin subdir como fallback.
        bytes_xlsx = None
        url_candidatas = [
            f"{BASE}/{pub_y}/{pub_m:02d}/{mes_es}_{y % 100:02d}.xlsx",
            f"{BASE}/{mes_es}_{y % 100:02d}.xlsx",
        ]
        for url in url_candidatas:
            try:
                resp = requests.get(url, timeout=10, headers={**HEADERS, "User-Agent": "Mozilla/5.0"})
                if resp.status_code == 200 and len(resp.content) > 5000 and resp.content[:2] == b"PK":
                    bytes_xlsx = resp.content
                    _record_last_modified("fiscal", resp)
                    break
            except Exception:
                continue
        if bytes_xlsx:
            if parse_y_mergear(bytes_xlsx, y, m):
                continue

        # Si no hay XLSX directo, intentar RAR del mes de publicación
        try:
            import rarfile
            for n in range(1, 10):
                url_rar = f"{BASE}/{pub_y}/{pub_m:02d}/cuentas_publicas{'' if n == 1 else n}.rar"
                try:
                    resp = requests.get(url_rar, timeout=15, headers={**HEADERS, "User-Agent": "Mozilla/5.0"})
                except Exception:
                    continue
                if resp.status_code != 200 or len(resp.content) < 5000:
                    continue
                # Es RAR válido?
                if resp.content[:4] not in (b"Rar!", b"\x52\x61\x72\x21"):
                    continue
                # Extraer XLSX
                tmp_path = SCRAPER_DIR / f"_tmp_fiscal_{y}_{m:02d}.rar"
                tmp_path.write_bytes(resp.content)
                try:
                    rf = rarfile.RarFile(str(tmp_path))
                    names = rf.namelist()
                    # Buscar XLSX que matchea el mes (ej. "Abril 26.xlsx")
                    target = None
                    mes_cap = mes_es.capitalize()
                    for n2 in names:
                        if n2.lower().endswith(".xlsx") and "imig" not in n2.lower() and mes_cap.lower() in n2.lower():
                            target = n2
                            break
                    if not target:
                        # Cualquier XLSX que no sea IMIG
                        for n2 in names:
                            if n2.lower().endswith(".xlsx") and "imig" not in n2.lower():
                                target = n2
                                break
                    if target:
                        xlsx_bytes = rf.read(target)
                        if parse_y_mergear(xlsx_bytes, y, m):
                            log(f"Fiscal Hacienda: {mes_es} {y} extraído de {url_rar.split('/')[-1]}/{target}", "ok")
                            break
                finally:
                    try:
                        tmp_path.unlink()
                    except Exception:
                        pass
        except ImportError:
            log("rarfile no instalado; saltando fiscal Hacienda RAR", "warn")
            break
        except Exception as e:
            log(f"Fiscal Hacienda RAR {mes_es} {y} falló: {e}", "warn")

    arr.sort(key=lambda x: x["f"])
    if nuevos > 0:
        log(f"Fiscal Hacienda: {nuevos} mes(es) nuevo(s), último {arr[-1]['f']}", "ok")
    return nuevos

def update_fiscal_indec(D: dict) -> int:
    """Resultado fiscal SPNF mensual desde la API de INDEC (IMIG).

    Más confiable que parsear los XLSX de Hacienda (que tienen fórmulas #REF! rotas).
    Schema: D.fiscal = [{f, sal_prim, sal_fin}] en MILLONES de pesos.
    """
    arr = D.setdefault("fiscal", [])
    nuevos = 0
    try:
        ids = [INDEC_SERIES["sal_primario"], INDEC_SERIES["sal_financiero"]]
        data = fetch_indec(ids, last=120)  # ~10 años
        for row in data:
            if not row[0]:
                continue
            f = ym(row[0])
            prim = row[1]
            fin = row[2]
            if prim is None and fin is None:
                continue
            rec_existing = next((x for x in arr if x.get("f") == f), None)
            new = {"f": f}
            if prim is not None:
                new["sal_prim"] = round(float(prim), 1)
            if fin is not None:
                new["sal_fin"] = round(float(fin), 1)
            if rec_existing is None:
                arr.append(new)
                nuevos += 1
            else:
                rec_existing.update({k: v for k, v in new.items() if k != "f"})
        arr.sort(key=lambda x: x["f"])
        if arr:
            log(f"Fiscal INDEC (IMIG): actualizado, último {arr[-1]['f']} (sal_prim={arr[-1].get('sal_prim','-')})", "ok")
    except Exception as e:
        log(f"Fiscal INDEC falló: {e}", "warn")
    return nuevos

def _parse_fiscal_xlsx(content: bytes) -> dict | None:
    """Extrae ingresos/gastos/sal_prim/sal_fin de un XLSX de Hacienda.

    Soporta dos formatos distintos del Esquema Ahorro-Inversión SPNF:

    Formato A — IMIG (preferido): hoja con nombre de mes en lowercase ('abril',
        'mayo'). Estructura: col 2 = CONCEPTO, col 7 = Dato mensual SPNF total.

    Formato B — Esquema Ahorro-Inversión clásico (filename mes_aa.xlsx):
        cols 3-6 = componentes Admin Nacional, col 7 = TOTAL Admin Nacional,
        col 8 = PAMI/Fdos Fiduciarios/Otros, col 9 = TOTAL SPNF.
        ⚠️ El TOTAL SPNF correcto está en COL 9, no en col 7.

    Formato C — VarMensual / SALIDA PRENSA: tiene #REF! por fórmulas rotas,
        se ignoran.

    Devuelve dict con valores en MILLONES de pesos, o None.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        return None

    KEYS = {
        "ingresos":   ["INGRESOS DESPUES DE FIGURAT", "INGRESOS TOTALES"],
        "gastos":     ["GASTOS DESPUES DE FIGURAT", "GASTOS TOTALES"],
        "sal_prim":   ["RESULTADO PRIMARIO", "SUPERAVIT PRIMARIO", "DEFICIT PRIMARIO"],
        "sal_fin":    ["RESULTADO FINANCIERO", "SUPERAVIT FINANCIERO", "DEFICIT FINANCIERO"],
    }
    MESES_LOW = {"enero","febrero","marzo","abril","mayo","junio","julio","agosto","septiembre","octubre","noviembre","diciembre"}

    def parse_hoja(ws, target_col):
        """Lee conceptos de una hoja usando la columna target_col como dato SPNF."""
        local = {"ingresos": None, "gastos": None, "sal_prim": None, "sal_fin": None}
        for r in range(1, ws.max_row + 1):
            concepto = ws.cell(r, 2).value
            if not concepto:
                continue
            cnorm = str(concepto).upper().strip()
            for key, patterns in KEYS.items():
                if local[key] is not None:
                    continue
                if any(p in cnorm for p in patterns):
                    v = ws.cell(r, target_col).value
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        local[key] = float(v)
                    break
        return local

    out = {"ingresos": None, "gastos": None, "sal_prim": None, "sal_fin": None}

    # Estrategia: priorizar formato A (IMIG hoja "abril") sobre B (Esquema)
    # Pero si solo está disponible Esquema, usar col 9 (no 7)
    # Skip hojas con #REF! conocido (VarMensual, SALIDA PRENSA, AIF)
    for sn in wb.sheetnames:
        if sn.lower() in ("varmensual", "aif"):
            continue
        if sn.upper().startswith("SALIDA PRENSA"):
            continue
        ws = wb[sn]
        if ws.max_row < 10:
            continue

        # Detectar formato
        sn_low = sn.strip().lower()
        is_imig_sheet = sn_low in MESES_LOW
        # Detectar Esquema: R1 contiene "SECRETARIA" o R5 contiene "ESQUEMA"
        r1 = str(ws.cell(1, 1).value or "")
        r5 = str(ws.cell(5, 1).value or "")
        is_esquema = ("SECRETARIA" in r1.upper()) or ("ESQUEMA" in r5.upper())

        # PRIORIDAD: si el archivo es Esquema (tiene "SECRETARIA" en R1), usar col 9
        # aunque la hoja se llame "Abril" (que confunde con formato IMIG).
        # Esto evita leer col 7 (Admin Nacional) en archivos Esquema.
        if is_esquema:
            target_col = 9  # TOTAL SPNF
        elif is_imig_sheet:
            target_col = 7  # IMIG: col 7 ya es el SPNF total
        else:
            continue

        local = parse_hoja(ws, target_col)
        # Mergear: priorizar valores ya encontrados, completar lo que falta
        for k, v in local.items():
            if out[k] is None and v is not None:
                out[k] = v
        if all(out[k] is not None for k in out):
            break

    if all(v is None for v in out.values()):
        return None
    return out

def update_fiscal(D: dict) -> int:
    """Sector Público Nacional No Financiero — Esquema Ahorro-Inversión base caja.

    Fuente: Secretaría de Hacienda (argentina.gob.ar/economia/sechacienda/infoestadistica).

    Patrones de URL observados:
      - `marzo_26.xlsx` (formato corto, sólo el mes más reciente)
      - `2026/02/sector_publico_base_caja_enero_2026.xlsx` (formato histórico)
    El archivo se publica en el MES SIGUIENTE al dato. Ej: enero 2026 → carpeta 2026/02/.

    Schema: D.fiscal = [{f, ingresos, gastos, sal_prim, sal_fin}] en millones de pesos.
    """
    MESES_ES = ["enero","febrero","marzo","abril","mayo","junio",
                "julio","agosto","septiembre","octubre","noviembre","diciembre"]
    BASE = "https://www.argentina.gob.ar/sites/default/files"
    arr = D.setdefault("fiscal", [])
    nuevos = 0
    today = datetime.now(timezone.utc)
    fallidos_seguidos = 0
    encontrados = []

    for delta in range(0, 96):  # hasta 8 años atrás (~2018)
        # Mes del DATO = today.month - 1 (el último publicado) - delta
        y, m = today.year, today.month - 1 - delta
        while m <= 0:
            m += 12
            y -= 1
        if y < 2017:
            break
        mes_es = MESES_ES[m - 1]
        # Mes de PUBLICACION = mes_dato + 1
        pub_y, pub_m = y, m + 1
        if pub_m > 12:
            pub_m -= 12
            pub_y += 1

        urls_try = [
            f"{BASE}/{mes_es}_{y % 100:02d}.xlsx",  # formato corto
            f"{BASE}/{pub_y}/{pub_m:02d}/sector_publico_base_caja_{mes_es}_{y}.xlsx",
            f"{BASE}/sector_publico_base_caja_{mes_es}_{y}.xlsx",
        ]
        bytes_xlsx = None
        for url in urls_try:
            try:
                resp = requests.get(url, timeout=15, headers={**HEADERS, "User-Agent": "Mozilla/5.0"})
                if resp.status_code == 200 and len(resp.content) > 5000 and resp.content[:2] == b"PK":
                    bytes_xlsx = resp.content
                    break
            except Exception:
                continue

        if bytes_xlsx is None:
            fallidos_seguidos += 1
            # Si llevamos muchos fallos seguidos cerca del fin de serie, cortar
            if fallidos_seguidos > 8 and len(encontrados) > 0:
                break
            continue
        fallidos_seguidos = 0

        parsed = _parse_fiscal_xlsx(bytes_xlsx)
        if parsed is None:
            continue

        f = f"{y:04d}-{m:02d}"
        rec_existing = next((x for x in arr if x.get("f") == f), None)
        new = {"f": f}
        for k, v in parsed.items():
            if v is not None:
                new[k] = round(v, 1)
        if rec_existing is None:
            arr.append(new)
            nuevos += 1
        else:
            rec_existing.update(new)
        encontrados.append(f)

    arr.sort(key=lambda x: x["f"])
    if arr:
        log(f"Fiscal SPNF: {len(encontrados)} meses procesados, total serie: {len(arr)}, último {arr[-1]['f']}", "ok")
    return nuevos

def find_ica_xls_url() -> str | None:
    """ICA publica `ica_cuadros_DD_MM_YY.xls` mensualmente. Buscar el más reciente."""
    today = datetime.now(timezone.utc)
    for delta in range(0, 5):
        y, m = today.year, today.month - delta
        while m <= 0:
            m += 12
            y -= 1
        for d in [20, 21, 22, 23, 19, 18, 17, 24, 25, 26, 27]:
            url = f"{INDEC_XLS_BASE}/ica_cuadros_{d:02d}_{m:02d}_{y % 100:02d}.xls"
            try:
                r = requests.head(url, timeout=8, headers={**HEADERS, "User-Agent": "Mozilla/5.0"})
                if r.status_code == 200 and "excel" in r.headers.get("Content-Type", "").lower():
                    return url
            except Exception:
                continue
    return None

def update_ica_xls(D: dict) -> int:
    """Intercambio Comercial Argentino — XLS oficial INDEC.

    Fuente: ica_cuadros_DD_MM_YY.xls Cuadro 1 (intercambio mensual del año en curso + anterior).
    Schema D.bc[i]: {f, expo, impo, impo_abs, saldo, pp, moa, moi, cye}.
    Esta función actualiza expo/impo/saldo. pp/moa/moi/cye vienen de la API series.
    """
    arr = D.setdefault("bc", [])
    nuevos = 0
    url = find_ica_xls_url()
    if not url:
        log("ICA XLS no encontrado (probados últimos 5 meses)", "warn")
        return 0
    try:
        wb = download_xls(url, serie="bc")
        sh = wb.sheet_by_name("c1")
        meses_idx = {"enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
                     "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12}
        # Detectar año actual/prev del header. INDEC cambió la fila del header
        # (antes R5, ahora R6: "Período | 2026e | 2025* | ..."). Escaneamos R4-R8
        # y nos quedamos con la fila que tenga el año en col 2.
        anio_actual = None
        anio_prev = None
        for hdr_row in range(4, 9):
            cand_act = cand_prev = None
            for c in range(2, sh.ncols):
                v = sh.cell_value(hdr_row, c)
                if isinstance(v, str):
                    digits = "".join(ch for ch in v if ch.isdigit())
                    if len(digits) == 4 and digits.startswith("20"):
                        if cand_act is None:
                            cand_act = int(digits)
                        elif cand_prev is None:
                            cand_prev = int(digits)
                            break
            # La fila válida tiene el año en la columna 2 (Expo año actual)
            if cand_act is not None:
                v2 = sh.cell_value(hdr_row, 2)
                if isinstance(v2, str) and "20" in "".join(ch for ch in v2 if ch.isdigit())[:2]:
                    anio_actual, anio_prev = cand_act, cand_prev
                    break
        if anio_actual is None:
            log("ICA XLS: no detecté año actual en header", "warn")
            return 0

        # Cuadro 1 layout:
        # Col 2 = Expo año actual, col 3 = Expo año prev
        # Col 6 = Impo año actual, col 7 = Impo año prev
        # Col 10 = Saldo año actual, col 11 = Saldo año prev
        # Filas con meses Enero-Diciembre (la fila exacta varía; detectamos por label).
        for row_idx in range(11, 26):
            mes_label = str(sh.cell_value(row_idx, 1)).strip().lower()
            if mes_label not in meses_idx:
                continue
            m = meses_idx[mes_label]
            for anio, col_expo, col_impo, col_saldo in [(anio_actual, 2, 6, 10), (anio_prev, 3, 7, 11)]:
                if anio is None:
                    continue
                expo = _safe_float(sh.cell_value(row_idx, col_expo))
                impo = _safe_float(sh.cell_value(row_idx, col_impo))
                saldo = _safe_float(sh.cell_value(row_idx, col_saldo))
                if expo is None and impo is None and saldo is None:
                    continue
                f = f"{anio:04d}-{m:02d}"
                rec = next((x for x in arr if x.get("f") == f), None)
                new = {"f": f}
                if expo is not None:
                    new["expo"] = round(expo, 1)
                if impo is not None:
                    new["impo_abs"] = round(impo, 1)
                    new["impo"] = -round(impo, 1)
                if saldo is not None:
                    new["saldo"] = round(saldo, 1)
                if rec is None:
                    arr.append(new)
                    nuevos += 1
                else:
                    for k, v in new.items():
                        if k != "f" and v is not None:
                            rec[k] = v
        arr.sort(key=lambda x: x["f"])
        # Último mes con expo válido
        ult_real = [r for r in arr if r.get("expo")]
        if ult_real:
            log(f"ICA (XLS oficial {url.split('/')[-1]}): actualizado, último {ult_real[-1]['f']}", "ok")
    except Exception as e:
        log(f"ICA XLS falló: {e}", "warn")
    return nuevos

def update_mayoristas(D: dict) -> int:
    """Inflación mayorista — IPIM, IPIB e IPP nivel general + breakdowns clave.

    Fuente: INDEC CSV oficial (indec.gob.ar/ftp/cuadros/economia/indice_*.csv)
      - IPIM: Índice de Precios Internos al por Mayor
      - IPIB: Índice de Precios Internos Básicos al por Mayor
      - IPP:  Índice de Precios Básicos del Productor

    Schema: D.ipim, D.ipib, D.ipp = [{f, n, vm, via}]. Cada serie tiene también
    breakdown por categoría guardado en .nac (nacionales), .imp (importados),
    .prim (primarios), .manuf (manufactura) cuando aplica.
    """
    BASE = "https://www.indec.gob.ar/ftp/cuadros/economia"
    fuentes = {"ipim": "indice_ipim.csv", "ipib": "indice_ipib.csv", "ipp": "indice_ipp.csv"}
    # Categorías que queremos extraer (la principal es ng_nivel_general)
    SUB_KEYS = {
        "ng_nivel_general":     "n",       # Nivel general
        "n_productos_nacionales": "nac",   # Productos nacionales
        "i_productos_importados": "imp",   # Productos importados
        "1_primarios":          "prim",    # Productos primarios
        "2_industria_manufacturera_ y_energia_electrica": "manuf",  # Manufactura+energía
    }

    total_nuevos = 0
    for clave, fn in fuentes.items():
        try:
            url = f"{BASE}/{fn}"
            r = http_get(url, serie=clave)
            text = r.content.decode("utf-8", errors="replace")
            lines = text.strip().split("\n")
            # Header: periodo;nivel_general_aperturas;indice_<clave>
            # Agrupar por mes
            por_mes = {}
            for line in lines[1:]:
                parts = line.strip().split(";")
                if len(parts) < 3:
                    continue
                periodo, categoria, valor_str = parts[0], parts[1].strip(), parts[2]
                if categoria not in SUB_KEYS:
                    continue
                try:
                    valor = float(valor_str.replace(",", "."))
                except ValueError:
                    continue
                f = periodo[:7]
                key_local = SUB_KEYS[categoria]
                por_mes.setdefault(f, {})[key_local] = round(valor, 2)

            arr = D.setdefault(clave, [])
            for f, vals in por_mes.items():
                rec = next((x for x in arr if x.get("f") == f), None)
                new = {"f": f, **vals}
                if rec is None:
                    arr.append(new)
                    total_nuevos += 1
                else:
                    rec.update({k: v for k, v in new.items() if k != "f"})
            arr.sort(key=lambda x: x["f"])
            # Calcular vm (mensual) y via (interanual) sobre n
            idx = {r["f"]: i for i, r in enumerate(arr)}
            for r in arr:
                if r.get("n") is None:
                    continue
                f = r["f"]
                prev_f = prev_month(f)
                py_f = prev_year(f)
                if prev_f in idx and arr[idx[prev_f]].get("n"):
                    r["vm"] = round((r["n"] / arr[idx[prev_f]]["n"] - 1) * 100, 2)
                if py_f in idx and arr[idx[py_f]].get("n"):
                    r["via"] = round((r["n"] / arr[idx[py_f]]["n"] - 1) * 100, 2)
            if arr:
                log(f"{clave.upper()}: actualizado (último {arr[-1]['f']}, n={arr[-1].get('n','—')} vm={arr[-1].get('vm','—')}%)", "ok")
        except Exception as e:
            log(f"{clave.upper()} CSV falló: {e}", "warn")
    return total_nuevos

def update_mora(D: dict) -> int:
    """Mora del sistema financiero (irregularidad de cartera) mensual.

    Fuente: BCRA Informe sobre Bancos · Anexo XLSX (InfBanc_Anexo.xlsx)
    Hoja: 'Calidad de Cartera (por líneas)'
    Filas: R59 = Familias - Cartera irregular total %
           R103 = Empresas - Cartera irregular total %

    Schema: D.mora = [{f, fam, emp}] en porcentaje.
    """
    arr = D.setdefault("mora", [])
    nuevos = 0
    url = "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/informes/InfBanc_Anexo.xlsx"
    try:
        import openpyxl
        r = http_get(url, serie="mora")
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
        # Buscar hoja con tilde o sin tilde
        sheet_name = None
        for sn in wb.sheetnames:
            if "Calidad de Cartera" in sn and "líneas" in sn.lower():
                sheet_name = sn
                break
            if "Calidad de Cartera" in sn and "lineas" in sn.lower():
                sheet_name = sn
                break
        if sheet_name is None:
            # Tomar la segunda hoja que contiene "Calidad" (por defecto)
            cands = [sn for sn in wb.sheetnames if "Calidad" in sn]
            sheet_name = cands[1] if len(cands) > 1 else (cands[0] if cands else None)
        if sheet_name is None:
            log("Mora BCRA: no encontré hoja Calidad de Cartera", "warn")
            return 0
        ws = wb[sheet_name]

        # Recorrer columnas desde col 2 (col 1 son labels)
        # R6 = fechas. R59 = familias %. R103 = empresas %.
        for c in range(2, ws.max_column + 1):
            fecha_cell = ws.cell(6, c).value
            fam_v = ws.cell(59, c).value
            emp_v = ws.cell(103, c).value
            if fecha_cell is None:
                continue
            if hasattr(fecha_cell, "strftime"):
                f = fecha_cell.strftime("%Y-%m")
            else:
                f = str(fecha_cell)[:7]
            if not f or not f.startswith("20"):
                continue
            fam_f = float(fam_v) if isinstance(fam_v, (int, float)) else None
            emp_f = float(emp_v) if isinstance(emp_v, (int, float)) else None
            if fam_f is None and emp_f is None:
                continue
            rec = next((x for x in arr if x.get("f") == f), None)
            new = {"f": f}
            if fam_f is not None:
                new["fam"] = round(fam_f, 2)
            if emp_f is not None:
                new["emp"] = round(emp_f, 2)
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                rec.update({k: v for k, v in new.items() if k != "f"})
                # Si el record era preliminar (prov), el dato oficial lo reemplaza
                if rec.pop("prov", False):
                    log(f"Mora {f}: dato oficial BCRA reemplaza al preliminar", "ok")
        arr.sort(key=lambda x: x["f"])
        if arr:
            log(f"Mora BCRA: actualizado (último {arr[-1]['f']}, fam={arr[-1].get('fam','—')}% emp={arr[-1].get('emp','—')}%)", "ok")
    except Exception as e:
        log(f"Mora BCRA falló: {e}", "warn")
    return nuevos

def update_merval(D: dict) -> int:
    """MERVAL via Yahoo Finance directo (sin proxy CORS porque corremos en GitHub Actions)."""
    arr = D.setdefault("merval", [])
    nuevos = 0
    try:
        url = "https://query1.finance.yahoo.com/v8/finance/chart/%5EMERV"
        data = get_json(url, params={"interval": "1d", "range": "5y"},
                        headers={**HEADERS, "User-Agent": "Mozilla/5.0"})
        result = data["chart"]["result"][0]
        ts = result["timestamp"]
        closes = result["indicators"]["quote"][0]["close"]
        # Agrupar por mes, último día
        by_month = {}
        for t, c in zip(ts, closes):
            if c is None:
                continue
            fecha = datetime.fromtimestamp(t, tz=timezone.utc).strftime("%Y-%m")
            by_month[fecha] = c  # se sobreescribe, queda el último
        for f, c in by_month.items():
            v = round(float(c) / 1000, 1)  # convertir a "miles de puntos" como el histórico
            rec = next((x for x in arr if x.get("f") == f), None)
            if rec is None:
                arr.append({"f": f, "v": v})
                nuevos += 1
            else:
                rec["v"] = v
        arr.sort(key=lambda x: x["f"])
        if by_month:
            ult = max(by_month.keys())
            log(f"MERVAL Yahoo: actualizado (último {ult} = {round(by_month[ult]/1000, 1)}K)", "ok")
    except Exception as e:
        log(f"MERVAL Yahoo falló: {e}", "warn")
    return nuevos

# ──────────────────────────────────────────────────────────────────────────────
# INDEC XLS oficiales (cuando la API series está atrasada vs los Excel publicados)
# ──────────────────────────────────────────────────────────────────────────────
MONTHS_ES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
INDEC_XLS_BASE = "https://www.indec.gob.ar/ftp/cuadros/economia"

# Fechas reales de publicacion capturadas del header HTTP Last-Modified de cada fuente.
# Se llena en cada descarga y consume desde calcular_fecha_publicacion().
SOURCE_LAST_MOD: dict[str, datetime] = {}

def _record_last_modified(serie: str, response: requests.Response) -> None:
    """Lee el header Last-Modified de un response y lo guarda en SOURCE_LAST_MOD[serie]."""
    if not serie:
        return
    lm = response.headers.get("Last-Modified")
    if not lm:
        return
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(lm)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        # Si ya hay una fecha, quedarse con la mas reciente (cada XLS por anio en IPI, etc.)
        prev = SOURCE_LAST_MOD.get(serie)
        if prev is None or dt > prev:
            SOURCE_LAST_MOD[serie] = dt
    except Exception:
        pass

def http_get(url: str, serie: str | None = None, **kwargs) -> requests.Response:
    """Wrapper de requests.get que captura Last-Modified por serie."""
    kwargs.setdefault("timeout", TIMEOUT)
    headers = {**HEADERS, "User-Agent": "Mozilla/5.0", **kwargs.pop("headers", {})}
    r = requests.get(url, headers=headers, **kwargs)
    r.raise_for_status()
    if serie:
        _record_last_modified(serie, r)
    return r

def download_xls(url: str, serie: str | None = None) -> xlrd.book.Book:
    """Descarga un .xls de INDEC y devuelve el workbook (captura Last-Modified si serie)."""
    r = http_get(url, serie=serie)
    return xlrd.open_workbook(file_contents=r.content)

def _month_num(cell_str: str) -> int | None:
    """'Enero' / 'Enero*' → 1, 'Febrero' → 2, etc."""
    s = str(cell_str).strip().rstrip("*").strip()
    return (MONTHS_ES.index(s) + 1) if s in MONTHS_ES else None

def _safe_float(v) -> float | None:
    try:
        if v == "" or v == "///" or v is None:
            return None
        return float(v)
    except (ValueError, TypeError):
        return None

def parse_emae_xls(wb: xlrd.book.Book) -> list[dict]:
    """sh_emae_mensual_base2004.xls → [{f, orig, via, dest, vm}]

    INDEC renombró la hoja de 'EMAE' a 'Tabla' alrededor de mayo 2026.
    Probamos ambos nombres; si no funciona, caemos al primer sheet por índice.
    """
    sh = None
    for nombre in ("EMAE", "Tabla"):
        try:
            sh = wb.sheet_by_name(nombre)
            break
        except xlrd.biffh.XLRDError:
            continue
    if sh is None:
        sh = wb.sheet_by_index(0)
    out, current_year = [], None
    for r in range(sh.nrows):
        y_cell = sh.cell_value(r, 0)
        m_cell = sh.cell_value(r, 1)
        # Año: float (2004.0) o str ("2026*")
        if isinstance(y_cell, float) and y_cell > 1990:
            current_year = int(y_cell)
        elif isinstance(y_cell, str):
            s = y_cell.strip().rstrip("*").strip()
            if s.isdigit() and 1990 < int(s) < 2100:
                current_year = int(s)
        if not current_year:
            continue
        month_num = _month_num(m_cell) if m_cell else None
        if not month_num:
            continue
        orig = _safe_float(sh.cell_value(r, 2))
        via  = _safe_float(sh.cell_value(r, 3))
        dest = _safe_float(sh.cell_value(r, 4))
        vm   = _safe_float(sh.cell_value(r, 5))
        if orig is None:
            continue
        out.append({
            "f": f"{current_year:04d}-{month_num:02d}",
            "orig": round(orig, 2),
            "via":  round(via, 2)  if via  is not None else None,
            "dest": round(dest, 2) if dest is not None else None,
            "vm":   round(vm, 2)   if vm   is not None else None,
        })
    return out

def parse_ipi_xls(wb: xlrd.book.Book) -> list[dict]:
    """sh_ipi_manufacturero_YYYY.xls (Cuadro 1) → [{f, orig, via}]"""
    try:
        sh = wb.sheet_by_name("Cuadro 1")
    except xlrd.biffh.XLRDError:
        sh = wb.sheet_by_index(1)
    out, current_year = [], None
    for r in range(sh.nrows):
        y_cell = sh.cell_value(r, 1)
        m_cell = sh.cell_value(r, 2)
        if isinstance(y_cell, float) and y_cell > 1990:
            current_year = int(y_cell)
        elif isinstance(y_cell, str):
            s = y_cell.strip().rstrip("*").strip()
            if s.isdigit() and 1990 < int(s) < 2100:
                current_year = int(s)
        if not current_year:
            continue
        month_num = _month_num(m_cell) if m_cell else None
        if not month_num:
            continue
        orig = _safe_float(sh.cell_value(r, 3))
        via  = _safe_float(sh.cell_value(r, 4))
        if orig is None:
            continue
        out.append({
            "f": f"{current_year:04d}-{month_num:02d}",
            "orig": round(orig, 2),
            "via":  round(via, 2) if via is not None else None,
        })
    return out

# Mapeo columnas UCI → keys de nuestro schema
UCI_COL_MAP = {
    1: "tot",     # Nivel general
    2: "alim",    # Productos alimenticios
    3: "tabaco",  # Productos del tabaco
    4: "text",    # Productos textiles
    5: "papel",
    6: "impr",    # Edición/impresión
    7: "refin",   # Refinación
    8: "quim",    # Sustancias químicas
    9: "caucho",
    10: "min",    # Productos minerales
    11: "metal",  # Industrias metálicas básicas
    12: "auto",   # Industria automotriz
    13: "metalm", # Metalmecánica
}

def parse_uci_xls(wb: xlrd.book.Book) -> list[dict]:
    """sh_capacidad_MM_YY.xls → [{f, tot, alim, ..., metalm}]"""
    sh = wb.sheet_by_index(0)
    out, current_year = [], None
    for r in range(sh.nrows):
        col0 = sh.cell_value(r, 0)
        if isinstance(col0, str):
            s = col0.strip()
            # "Año 2026" o "Año 2025*"
            if s.startswith("Año "):
                ystr = s[4:].strip().rstrip("*").strip()
                if ystr.isdigit():
                    current_year = int(ystr)
                continue
            month_num = _month_num(s)
            if not month_num or not current_year:
                continue
            rec = {"f": f"{current_year:04d}-{month_num:02d}"}
            for col, key in UCI_COL_MAP.items():
                v = _safe_float(sh.cell_value(r, col))
                if v is not None:
                    rec[key] = round(v, 1)
            if "tot" in rec:
                out.append(rec)
    return out

def find_uci_xls_url() -> str | None:
    """sh_capacidad_MM_YY.xls — prueba del mes actual hacia atrás (máx 6 meses)."""
    today = datetime.now(timezone.utc)
    for delta in range(0, 7):
        y, m = today.year, today.month - delta
        while m <= 0:
            m += 12
            y -= 1
        url = f"{INDEC_XLS_BASE}/sh_capacidad_{m:02d}_{y % 100:02d}.xls"
        try:
            r = requests.head(url, timeout=10, headers={**HEADERS, "User-Agent": "Mozilla/5.0"})
            if r.status_code == 200 and "excel" in r.headers.get("Content-Type", "").lower():
                return url
        except Exception:
            continue
    return None

# Mapeo EMAE actividad: cols del XLS → keys del schema bedrock D.emae[].s
EMAE_S_COLS = {
    2:  "Agro",      # A - Agricultura
    3:  "Pesca",     # B - Pesca
    4:  "Mineria",   # C - Minería
    5:  "Industria", # D - Industria manufacturera
    6:  "EGA",       # E - Electricidad/Gas/Agua
    7:  "Const",     # F - Construcción
    8:  "Comer",     # G - Comercio
    9:  "Hoteles",   # H - Hoteles y restaurantes
    10: "Transp",    # I - Transporte y comunicaciones
    11: "Financ",    # J - Intermediación financiera
    12: "Inmob",     # K - Actividades inmobiliarias
    13: "AdmPub",    # L - Administración pública
    14: "Ensen",     # M - Enseñanza
    15: "Salud_s",   # N - Salud
}

def parse_emae_actividad_xls(wb: xlrd.book.Book) -> list[dict]:
    """sh_emae_actividad_base2004.xls (hoja 'Tabla Letras') → [{f, s:{Agro,Pesca,...}}]"""
    sh = None
    for nombre in ("Tabla Letras", "Tabla", "Letras"):
        try:
            sh = wb.sheet_by_name(nombre)
            break
        except xlrd.biffh.XLRDError:
            continue
    if sh is None:
        sh = wb.sheet_by_index(0)
    out, current_year = [], None
    for r in range(sh.nrows):
        y_cell = sh.cell_value(r, 0)
        m_cell = sh.cell_value(r, 1)
        if isinstance(y_cell, float) and y_cell > 1990:
            current_year = int(y_cell)
        elif isinstance(y_cell, str):
            s = y_cell.strip().rstrip("*").strip()
            if s.isdigit() and 1990 < int(s) < 2100:
                current_year = int(s)
        if not current_year:
            continue
        month_num = _month_num(m_cell) if m_cell else None
        if not month_num:
            continue
        sectores = {}
        for col, name in EMAE_S_COLS.items():
            v = _safe_float(sh.cell_value(r, col))
            if v is not None:
                sectores[name] = round(v, 2)
        if sectores:
            out.append({"f": f"{current_year:04d}-{month_num:02d}", "s": sectores})
    return out

# Mapeo IPI Cuadro 5: cols del XLS (impares = nivel del sector) → keys de D.ipi[].r
IPI_R_COLS = {
    5:  "Alimentos",
    7:  "Tabaco",
    9:  "Textiles",
    11: "Vestim",
    13: "Madera",
    15: "Petroleo",
    17: "Quimica",
    19: "Caucho",
    21: "Minerales",
    23: "MetalBase",
    25: "ProdMetal",
    27: "Maquinaria",
    29: "OtrosEq",
    31: "Automotriz",
    33: "OtroTransp",
    35: "Muebles",
}

def parse_ipi_sectores_xls(wb: xlrd.book.Book) -> list[dict]:
    """sh_ipi_manufacturero_YYYY.xls (Cuadro 5) → [{f, r:{Alimentos,Tabaco,...}}]"""
    try:
        sh = wb.sheet_by_name("Cuadro 5")
    except xlrd.biffh.XLRDError:
        return []
    out, current_year = [], None
    for r in range(sh.nrows):
        y_cell = sh.cell_value(r, 1)
        m_cell = sh.cell_value(r, 2)
        if isinstance(y_cell, float) and y_cell > 1990:
            current_year = int(y_cell)
        elif isinstance(y_cell, str):
            s = y_cell.strip().rstrip("*").strip()
            if s.isdigit() and 1990 < int(s) < 2100:
                current_year = int(s)
        if not current_year:
            continue
        month_num = _month_num(m_cell) if m_cell else None
        if not month_num:
            continue
        ramas = {}
        for col, name in IPI_R_COLS.items():
            v = _safe_float(sh.cell_value(r, col))
            if v is not None:
                ramas[name] = round(v, 2)
        if ramas:
            out.append({"f": f"{current_year:04d}-{month_num:02d}", "r": ramas})
    return out

def update_emae_sectores_xls(D: dict) -> int:
    """Completa D.emae[].s desde el XLS de EMAE actividad."""
    arr = D.setdefault("emae", [])
    nuevos = 0
    try:
        wb = download_xls(f"{INDEC_XLS_BASE}/sh_emae_actividad_base2004.xls", serie="emae")
        rows = parse_emae_actividad_xls(wb)
        for new in rows:
            rec = next((x for x in arr if x.get("f") == new["f"]), None)
            if rec is None:
                # No deberíamos llegar acá si update_emae_xls corrió antes, pero por las dudas
                arr.append(new)
                nuevos += 1
            else:
                rec["s"] = new["s"]
        if rows:
            log(f"EMAE sectores (XLS oficial): actualizado (último {rows[-1]['f']})", "ok")
    except Exception as e:
        log(f"EMAE sectores XLS falló: {e}", "warn")
    return nuevos

def update_ipi_sectores_xls(D: dict) -> int:
    """Completa D.ipi[].r desde el Cuadro 5 del XLS de IPI."""
    arr = D.setdefault("ipi", [])
    nuevos = 0
    today = datetime.now(timezone.utc)
    wb = None
    for year in (today.year, today.year - 1):
        try:
            wb = download_xls(f"{INDEC_XLS_BASE}/sh_ipi_manufacturero_{year}.xls", serie="ipi")
            break
        except Exception:
            continue
    if wb is None:
        return 0
    try:
        rows = parse_ipi_sectores_xls(wb)
        for new in rows:
            rec = next((x for x in arr if x.get("f") == new["f"]), None)
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                rec["r"] = new["r"]
        if rows:
            log(f"IPI sectores (XLS Cuadro 5): actualizado (último {rows[-1]['f']})", "ok")
    except Exception as e:
        log(f"IPI sectores XLS falló: {e}", "warn")
    return nuevos

def update_emae_xls(D: dict) -> int:
    """EMAE oficial desde el XLS de INDEC (más actualizado que la API series)."""
    arr = D.setdefault("emae", [])
    nuevos = 0
    try:
        wb = download_xls(f"{INDEC_XLS_BASE}/sh_emae_mensual_base2004.xls", serie="emae")
        rows = parse_emae_xls(wb)
        for new in rows:
            rec = next((x for x in arr if x.get("f") == new["f"]), None)
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                # No pisar la subdivisión sectorial 's' si la tiene del bedrock
                for k, v in new.items():
                    if v is not None and k != "s":
                        rec[k] = v
        arr.sort(key=lambda x: x.get("f", ""))
        if rows:
            log(f"EMAE (XLS oficial): actualizado (último {rows[-1]['f']})", "ok")
    except Exception as e:
        log(f"EMAE XLS falló: {e}", "warn")
    return nuevos

def update_ipi_xls(D: dict) -> int:
    """IPI oficial desde el XLS de INDEC. Prueba YYYY actual y anterior si 404."""
    arr = D.setdefault("ipi", [])
    nuevos = 0
    today = datetime.now(timezone.utc)
    wb = None
    for year in (today.year, today.year - 1):
        url = f"{INDEC_XLS_BASE}/sh_ipi_manufacturero_{year}.xls"
        try:
            wb = download_xls(url, serie="ipi")
            break
        except Exception:
            continue
    if wb is None:
        log("IPI XLS no encontrado", "warn")
        return 0
    try:
        rows = parse_ipi_xls(wb)
        for new in rows:
            rec = next((x for x in arr if x.get("f") == new["f"]), None)
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                for k, v in new.items():
                    if v is not None:
                        rec[k] = v
        arr.sort(key=lambda x: x.get("f", ""))
        # Recalcular vm (mensual sobre dest) si tenemos serie desestacionalizada
        for i, r in enumerate(arr):
            if i > 0 and r.get("dest") and arr[i-1].get("dest"):
                r["vm"] = round((r["dest"] / arr[i-1]["dest"] - 1) * 100, 2)
        if rows:
            log(f"IPI (XLS oficial): actualizado (último {rows[-1]['f']})", "ok")
    except Exception as e:
        log(f"IPI XLS parse falló: {e}", "warn")
    return nuevos

def update_uci_xls(D: dict) -> int:
    """UCI Nivel General + sectorial desde el XLS oficial de INDEC."""
    arr = D.setdefault("uci", [])
    nuevos = 0
    url = find_uci_xls_url()
    if not url:
        log("UCI XLS no encontrado (probados últimos 6 meses)", "warn")
        return 0
    try:
        wb = download_xls(url, serie="uci")
        rows = parse_uci_xls(wb)
        for new in rows:
            rec = next((x for x in arr if x.get("f") == new["f"]), None)
            if rec is None:
                arr.append(new)
                nuevos += 1
            else:
                rec.update({k: v for k, v in new.items() if v is not None})
        arr.sort(key=lambda x: x.get("f", ""))
        if rows:
            log(f"UCI (XLS oficial {url.split('/')[-1]}): actualizado (último {rows[-1]['f']})", "ok")
    except Exception as e:
        log(f"UCI XLS parse falló: {e}", "warn")
    return nuevos

# ──────────────────────────────────────────────────────────────────────────────
# Detección de publicaciones nuevas (para sección "Esta semana" del dashboard)
# ──────────────────────────────────────────────────────────────────────────────
SERIES_TRACK_PUB = {
    # serie_key: (fuente_label, key del registro de fecha)
    "ipc":      ("INDEC", "f"),
    "ipim":     ("INDEC", "f"),
    "ipib":     ("INDEC", "f"),
    "ipp":      ("INDEC", "f"),
    "super":    ("INDEC", "f"),
    "mayor":    ("INDEC", "f"),
    "emae":     ("INDEC", "f"),
    "ipi":      ("INDEC", "f"),
    "isac":     ("INDEC", "f"),
    "uci":      ("INDEC", "f"),
    "bc":       ("INDEC", "f"),
    "turismo":  ("INDEC", "f"),
    "salarios": ("INDEC", "f"),
    "ripte":    ("INDEC", "f"),
    "trabajo":  ("INDEC (SIPA)", "f"),
    "empresas": ("OEDE", "a"),
    "rec":      ("AFIP/ARCA", "f"),
    "fiscal":   ("Hacienda (vía INDEC IMIG)", "f"),
    "mora":     ("BCRA Informe sobre Bancos", "f"),
    "eph":      ("INDEC EPH", "p"),
    "pbi":      ("INDEC CCNN", "p"),
}

SERIE_LBL = {
    "ipc":      "IPC inflación",
    "ipim":     "IPIM mayorista",
    "ipib":     "IPIB básicos",
    "ipp":      "IPP productor",
    "super":    "Encuesta Supermercados",
    "mayor":    "Encuesta Autoservicios Mayoristas",
    "emae":     "EMAE actividad",
    "ipi":      "IPI manufacturero",
    "isac":     "ISAC construcción",
    "uci":      "UCI capacidad instalada",
    "bc":       "ICA balanza comercial",
    "turismo":  "Turismo internacional",
    "salarios": "Salarios IS",
    "ripte":    "RIPTE",
    "trabajo":  "Empleo SIPA",
    "empresas": "Empresas activas",
    "rec":      "Recaudación tributaria",
    "fiscal":   "Resultado fiscal SPNF",
    "mora":     "Irregularidad cartera",
    "eph":      "Desocupación EPH",
    "pbi":      "PBI trimestral",
}

def _ultimo_real(arr: list, fecha_key: str) -> dict | None:
    """Último registro de una serie ignorando proyecciones."""
    if not isinstance(arr, list):
        return None
    candidatos = [r for r in arr if not r.get("proj") and r.get(fecha_key) is not None]
    if not candidatos:
        return None
    # Ordenar por período normalizado: 'YYYY-MM' y 'YYYY' funcionan como string,
    # pero los trimestres ('4to t. 2025' vs '1er t. 2026') requieren normalizar a
    # 'YYYY-Qn' para no ordenar alfabéticamente (4>1 daría Q4-2025 > Q1-2026).
    candidatos.sort(key=lambda r: _norm_periodo(str(r.get(fecha_key))))
    return candidatos[-1]

# Heurística: días después del FIN de período en que se publica oficialmente la serie
# Basado en el calendario habitual de difusión INDEC/BCRA/Hacienda
PUB_LAG_DIAS = {
    "ipc":      13,   # ~13 del mes siguiente al dato
    "ipim":     19,   # ~19 del mes siguiente (confirmado por INDEC: IPIM abr-26 publicado 19-may-26)
    "ipib":     19,
    "ipp":      19,
    "ipi":      24,
    "isac":     24,
    "uci":      24,
    "super":    52,   # Supermercados: ~52 días (marzo 2026 publicado 22-may-26)
    "mayor":    52,
    "bc":       21,   # ICA: 17-21 del mes siguiente
    "emae":     55,   # EMAE con lag de ~2 meses post-cierre
    "turismo":  26,   # Turismo internacional: ~26 días (abril 2026 publicado 26-may-26)
    "salarios": 50,
    "ripte":    35,
    "trabajo":  60,   # SIPA con ~2 meses lag
    "empresas": 70,   # OEDE anual, datos definitivos en feb/mar año siguiente
    "rec":      2,    # AFIP publica al día siguiente
    "fiscal":   22,   # Hacienda IMIG
    "mora":     28,   # BCRA Informe sobre Bancos
    "eph":      80,   # EPH trimestral, ~3 meses después del cierre
    "pbi":      90,   # CCNN trimestral
}

def _norm_periodo(p: str) -> str:
    """Normaliza períodos a forma canónica para comparar entre fuentes.

    'YYYY-MM' → igual. Trimestres ('1er t. 2026', '1°T 2026', '1T 2026',
    'primer trimestre 2026') → 'YYYY-Qn'. Años sueltos → 'YYYY'.
    """
    import re
    p = (p or "").strip().lower()
    if re.match(r"^\d{4}-\d{2}$", p):
        return p
    m = re.search(r"(\d{4})", p)
    if not m:
        return p
    y = m.group(1)
    qmap = {"1er": 1, "1°": 1, "1": 1, "primer": 1,
            "2do": 2, "2°": 2, "2": 2, "segundo": 2,
            "3er": 3, "3°": 3, "3": 3, "tercer": 3,
            "4to": 4, "4°": 4, "4": 4, "cuarto": 4}
    mq = re.search(r"(1er|2do|3er|4to|primer|segundo|tercer|cuarto|[1-4]°|[1-4])\s*(?:t\.?|trim|°t)", p)
    if mq and ("t" in p or "trim" in p or "°" in p):
        q = qmap.get(mq.group(1))
        if q:
            return f"{y}-Q{q}"
    return p

_CAL_CACHE: list | None = None
def _calendar_oficial_lookup(serie: str, periodo: str) -> datetime | None:
    """Busca un evento en calendar_oficial.json que matchee (serie_key, periodo).

    Devuelve datetime de publicación si existe, o None.
    """
    global _CAL_CACHE
    if _CAL_CACHE is None:
        path = SCRAPER_DIR / "calendar_oficial.json"
        if not path.exists():
            _CAL_CACHE = []
        else:
            try:
                _CAL_CACHE = json.loads(path.read_text(encoding="utf-8")).get("eventos", [])
            except Exception:
                _CAL_CACHE = []
    if not _CAL_CACHE:
        return None
    periodo_norm = _norm_periodo(periodo)
    for ev in _CAL_CACHE:
        if ev.get("serie_key") != serie:
            continue
        if _norm_periodo(ev.get("periodo", "")) != periodo_norm:
            continue
        # Solo confiamos en calendarios OFICIALES (INDEC PDF, BCRA HTML). Las
        # entradas de Hacienda/ARCA son estimaciones propias — para esas usamos
        # el Last-Modified real del archivo (paso 2 de calcular_fecha_publicacion).
        if ev.get("fuente") in ("Hacienda", "ARCA"):
            return None
        try:
            return datetime.fromisoformat(ev["fecha"]).replace(tzinfo=timezone.utc)
        except Exception:
            continue
    return None

def calcular_fecha_publicacion(serie: str, periodo: str) -> datetime:
    """Devuelve cuándo se publicó oficialmente un período de una serie.

    Prioridad:
      1) Calendario oficial (calendar_oficial.json): fecha exacta del PDF INDEC
         o HTML BCRA para esta (serie, periodo). Es la fuente de verdad.
      2) SOURCE_LAST_MOD[serie]: header HTTP Last-Modified real del archivo
         descargado en este run (cuando hay XLSX).
      3) Heurística PUB_LAG_DIAS como fallback.
    """
    from calendar import monthrange
    import re
    today = datetime.now(timezone.utc)
    # 1) Calendario oficial
    cal = _calendar_oficial_lookup(serie, periodo)
    if cal is not None:
        return cal
    # 2) Last-Modified real del archivo fuente
    real = SOURCE_LAST_MOD.get(serie)
    if real is not None:
        return real
    lag = PUB_LAG_DIAS.get(serie, 30)
    try:
        if len(periodo) == 4 and periodo.isdigit():
            # Anual: fin = 31 dic
            fin = datetime(int(periodo), 12, 31, tzinfo=timezone.utc)
        elif "-" in periodo and len(periodo) >= 7:
            # YYYY-MM
            y, m = int(periodo[:4]), int(periodo[5:7])
            last_day = monthrange(y, m)[1]
            fin = datetime(y, m, last_day, tzinfo=timezone.utc)
        elif "trim" in periodo.lower() or "t." in periodo.lower():
            # "1er trimestre 2025" o "1er t. 2025"
            qm = re.search(r"(\d+)(?:er|do|to|°|º)?\s*t", periodo.lower())
            ym = re.search(r"(\d{4})", periodo)
            if qm and ym:
                trimestre = int(qm.group(1))
                year = int(ym.group(1))
                mes_fin = min(trimestre * 3, 12)
                last_day = monthrange(year, mes_fin)[1]
                fin = datetime(year, mes_fin, last_day, tzinfo=timezone.utc)
            else:
                return today
        else:
            return today
    except Exception:
        return today
    publicacion = fin + timedelta(days=lag)
    if publicacion > today:
        publicacion = today
    return publicacion

def detect_publicaciones(D_old: dict, D_new: dict) -> list[dict]:
    """Detecta publicaciones recientes con fecha estimada según calendario INDEC/BCRA.

    Reglas:
      - Para cada serie, toma su último período real (no proyección).
      - Calcula `publicado_at` aplicando PUB_LAG_DIAS al fin del período.
      - Si `publicado_at` cae en el futuro, clamp a hoy.
      - Si `publicado_at` está dentro de los últimos 30 días → registra.
      - Sin duplicar: una pub por (serie, periodo) única en la lista acumulada.
    """
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(days=30)
    cutoff_iso = cutoff.isoformat()

    # Recuperar lista existente del data.json viejo
    pubs = (D_old.get("_meta") or {}).get("publicaciones") or []
    # Re-calcular publicado_at Y refrescar datos de las pubs existentes en cada run.
    # publicado_at: importante si ajustamos PUB_LAG_DIAS.
    # datos: importante si el parser corrigió valores (ej. fiscal SPNF col 9 vs col 7).
    for p in pubs:
        serie_key = p.get("serie")
        periodo = str(p.get("periodo", ""))
        if not serie_key or not periodo:
            continue
        # Recalcular fecha de publicación (priorizar _pub_at del record si existe)
        try:
            fecha_key0 = SERIES_TRACK_PUB.get(serie_key, (None, None))[1]
            new_arr0 = D_new.get(serie_key) or []
            rec0 = next((r for r in new_arr0 if str(r.get(fecha_key0, "")) == periodo), None) if fecha_key0 else None
            pub_ov = rec0.get("_pub_at") if rec0 else None
            if pub_ov:
                p["publicado_at"] = pub_ov
            else:
                fecha_pub = calcular_fecha_publicacion(serie_key, periodo)
                p["publicado_at"] = fecha_pub.isoformat(timespec="seconds")
        except Exception:
            pass
        # Refrescar datos desde D_new[serie][periodo] (snapshot actualizado)
        fecha_key = SERIES_TRACK_PUB.get(serie_key, (None, None))[1]
        if not fecha_key:
            continue
        new_arr = D_new.get(serie_key)
        if not isinstance(new_arr, list):
            continue
        rec = next((r for r in new_arr if str(r.get(fecha_key, "")) == periodo), None)
        if rec:
            vals = {k: v for k, v in rec.items() if k not in (fecha_key, "proj", "s", "r", "comp")}
            datos = {k: round(v, 2) if isinstance(v, float) else v
                     for k, v in vals.items() if isinstance(v, (int, float)) and not isinstance(v, bool)}
            if datos:
                p["datos"] = datos
    # Filtrar las viejas (>30d) tras el recálculo
    pubs = [p for p in pubs if p.get("publicado_at", p.get("detectado_at", "")) > cutoff_iso]

    for serie_key, (fuente, fecha_key) in SERIES_TRACK_PUB.items():
        new_arr = D_new.get(serie_key)
        if not isinstance(new_arr, list) or not new_arr:
            continue
        ult_new = _ultimo_real(new_arr, fecha_key)
        if ult_new is None:
            continue
        per_new = str(ult_new.get(fecha_key, ""))
        if not per_new:
            continue

        # Calcular fecha de publicación: si el record tiene _pub_at (override manual), usarla;
        # sino heurística Last-Modified / lag.
        pub_override = ult_new.get("_pub_at")
        if pub_override:
            try:
                fecha_pub = datetime.fromisoformat(pub_override)
            except Exception:
                fecha_pub = calcular_fecha_publicacion(serie_key, per_new)
        else:
            fecha_pub = calcular_fecha_publicacion(serie_key, per_new)
        # Solo registrar si publicado_at está dentro de los últimos 30 días
        if fecha_pub < cutoff:
            continue

        # No duplicar la misma (serie, periodo) ya en pubs
        if any(p.get("serie") == serie_key and p.get("periodo") == per_new for p in pubs):
            continue

        vals = {kk: vv for kk, vv in ult_new.items() if kk not in (fecha_key, "proj", "s", "r", "comp")}
        datos = {kk: round(vv, 2) if isinstance(vv, float) else vv
                 for kk, vv in vals.items() if isinstance(vv, (int, float)) and not isinstance(vv, bool)}
        pubs.append({
            "serie": serie_key,
            "label": SERIE_LBL.get(serie_key, serie_key),
            "fuente": fuente,
            "periodo": per_new,
            "publicado_at": fecha_pub.isoformat(timespec="seconds"),
            "detectado_at": now.isoformat(timespec="seconds"),
            "datos": datos,
        })

    # Ordenar por publicado_at más reciente primero
    pubs.sort(key=lambda p: p.get("publicado_at", p.get("detectado_at", "")), reverse=True)
    return pubs

# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────
# Calendario forward looking: próximas publicaciones esperadas
# ──────────────────────────────────────────────────────────────────────────────
# Series de actualización diaria — no aparecen en el calendario semanal.
SERIES_DIARIAS = {"tc", "reservas", "embi", "merval", "bm"}

# Periodicidad de cada serie. Las anuales/trimestrales se manejan distinto a las
# mensuales para calcular el "siguiente periodo esperado".
SERIE_PERIODICIDAD = {
    "empresas": "anual",
    "eph":      "trimestral",
    "pbi":      "trimestral",
    # resto: mensual por default
}

def _siguiente_periodo(periodo: str, periodicidad: str = "mensual") -> str:
    """Dado un período (YYYY-MM, YYYY, '1T 2026'), devuelve el siguiente."""
    if periodicidad == "anual":
        try:
            return str(int(periodo) + 1)
        except Exception:
            return periodo
    if periodicidad == "trimestral":
        import re
        m = re.match(r"(\d)[°ºT]\s*(\d{4})", periodo) or re.match(r"(\d{4})-Q(\d)", periodo)
        if m:
            if "-Q" in periodo:
                y, q = int(m.group(1)), int(m.group(2))
            else:
                q, y = int(m.group(1)), int(m.group(2))
            nq = q + 1; ny = y
            if nq > 4: nq = 1; ny += 1
            return f"{nq}°T {ny}"
        return periodo
    # mensual: YYYY-MM
    try:
        y, m = int(periodo[:4]), int(periodo[5:7])
        m += 1
        if m > 12: m = 1; y += 1
        return f"{y:04d}-{m:02d}"
    except Exception:
        return periodo

def _fin_periodo(periodo: str, periodicidad: str = "mensual") -> datetime:
    """Devuelve el último día del período (datetime UTC)."""
    from calendar import monthrange
    if periodicidad == "anual":
        return datetime(int(periodo), 12, 31, tzinfo=timezone.utc)
    if periodicidad == "trimestral":
        import re
        m = re.match(r"(\d)[°ºT]\s*(\d{4})", periodo)
        if m:
            q, y = int(m.group(1)), int(m.group(2))
            mes = q * 3
            return datetime(y, mes, monthrange(y, mes)[1], tzinfo=timezone.utc)
        return datetime.now(timezone.utc)
    y, m = int(periodo[:4]), int(periodo[5:7])
    return datetime(y, m, monthrange(y, m)[1], tzinfo=timezone.utc)

def compute_proximas_publicaciones(D: dict, days_ahead: int = 7) -> list[dict]:
    """Devuelve eventos de publicación esperados en los próximos N días.

    Lee scraper/calendar_oficial.json (generado por fetch_calendars.py que parsea
    los PDFs/HTML oficiales de INDEC, BCRA, ARCA, Hacienda). NO estima fechas:
    expone solo lo que los organismos publicaron en sus calendarios.
    """
    cal_path = SCRAPER_DIR / "calendar_oficial.json"
    if not cal_path.exists():
        log("calendar_oficial.json no existe — corré 'python scraper/fetch_calendars.py' primero", "warn")
        return []
    try:
        cal = json.loads(cal_path.read_text(encoding="utf-8"))
    except Exception as e:
        log(f"No pude leer calendar_oficial.json: {e}", "warn")
        return []
    hoy = datetime.now(timezone.utc).date()
    cutoff = hoy + timedelta(days=days_ahead)
    proximas = []
    for ev in cal.get("eventos", []):
        try:
            fecha = datetime.fromisoformat(ev["fecha"]).date()
        except Exception:
            continue
        if not (hoy <= fecha <= cutoff):
            continue
        serie = ev.get("serie_key")
        # Marcar si la serie ya tiene un dato preliminar/prov esperando oficial
        prov_pendiente = False
        if serie and isinstance(D.get(serie), list):
            fecha_key = SERIES_TRACK_PUB.get(serie, (None, "f"))[1] or "f"
            ult = _ultimo_real(D[serie], fecha_key)
            if ult and ult.get("prov"):
                prov_pendiente = True
        proximas.append({
            "serie": serie,
            "label": ev.get("label", ""),
            "fuente": ev.get("fuente", ""),
            "periodo_esperado": ev.get("periodo", ""),
            "fecha_estimada": ev["fecha"],
            "prov_pendiente": prov_pendiente,
        })
    return proximas

# ──────────────────────────────────────────────────────────────────────────────
# Overrides manuales: datos preliminares de notas / consultoras
# Cada entrada se aplica SOLO si el periodo no existe ya en la serie (no pisa
# al dato oficial cuando llega). Llevan flag prov:True + fuente:str.
# ──────────────────────────────────────────────────────────────────────────────
MANUAL_OVERRIDES = [
    {
        "serie": "mora", "f": "2026-04",
        "data": {"fam": 12.0, "emp": 3.3},
        "fuente": "Infobae/1816 via CENDEU",
        "publicado_at": "2026-06-02T00:00:00+00:00",
    },
]

def apply_manual_overrides(D: dict) -> None:
    for ov in MANUAL_OVERRIDES:
        arr = D.setdefault(ov["serie"], [])
        if any(x.get("f") == ov["f"] for x in arr):
            continue  # ya existe el periodo (oficial o previamente cargado)
        rec = {"f": ov["f"], **ov["data"], "prov": True, "fuente": ov["fuente"], "_pub_at": ov["publicado_at"]}
        arr.append(rec)
        arr.sort(key=lambda x: x.get("f", ""))
        log(f"Override manual: {ov['serie']} {ov['f']} (prov, {ov['fuente']})", "ok")

# ──────────────────────────────────────────────────────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="No escribe data.json")
    args = ap.parse_args()

    log(f"Cargando bedrock desde {HISTORICAL.name}")
    D = json.loads(HISTORICAL.read_text(encoding="utf-8"))
    log(f"Bedrock: {len(D)} series, {sum(len(v) for v in D.values() if isinstance(v, list))} registros")

    # Cargar data.json previo (resultado del run anterior) para:
    #   1) Detectar publicaciones nuevas comparando.
    #   2) ACTUAR COMO CONTINUIDAD para series que no estan en historical.json
    #      (ej. mora: bedrock vacio, todo viene del XLSX BCRA. Si el fetch falla,
    #      perderiamos la historia sin esta proteccion).
    D_prev = {}
    if OUT_PATH.exists():
        try:
            D_prev = json.loads(OUT_PATH.read_text(encoding="utf-8"))
            log(f"Cargado data.json previo para detección de publicaciones nuevas")
            # Continuidad: para cada serie del previo, si el bedrock no la tiene
            # (o la tiene vacia) y el previo si, usar el previo como punto de partida.
            # Esto protege contra fetch fallidos: si update_X no agrega registros,
            # mantenemos la historia anterior en lugar de quedar con un array vacio.
            restored = []
            for k, v in D_prev.items():
                if k == "_meta" or not isinstance(v, list) or not v:
                    continue
                if not D.get(k):  # ausente o []
                    D[k] = [dict(r) for r in v]  # deep copy de records
                    restored.append(f"{k}({len(v)})")
            if restored:
                log(f"Continuidad desde data.json previo: {', '.join(restored)}", "ok")
        except Exception as e:
            log(f"No se pudo leer data.json previo: {e}", "warn")

    # Actualizar cada bloque (cada uno maneja sus propios errores)
    update_ipc(D)
    update_ipc_divisiones(D)  # desglose COICOP oficial (CSV, mismo dia que el IPC)
    # EMAE / IPI / UCI: primero API series (rápido), después XLS oficial (más fresco — pisa la API)
    update_emae(D)
    update_emae_xls(D)
    update_emae_sectores_xls(D)
    update_ipi_isac(D)
    update_ipi_xls(D)
    update_ipi_sectores_xls(D)
    update_uci_xls(D)
    update_bcra_monthly(D)
    update_embi(D)
    update_simple_monthly(D, "ripte", INDEC_SERIES["ripte"], "nom", "RIPTE")
    update_simple_monthly(D, "rec",   INDEC_SERIES["rec"],   "tot", "Recaudación")
    update_uci_sectorial(D)
    update_salarios(D)
    update_salarios_csv(D)  # complementa: CSV oficial más fresco que la API series
    update_comercio_interior(D)  # Supermercados + Autoservicios mayoristas
    update_empleo(D)
    update_eph(D)
    update_empresas(D)
    update_fiscal_indec(D)
    update_fiscal_hacienda(D)  # complementa: fiscal Hacienda llega antes que IMIG INDEC
    update_mora(D)
    update_ica_xls(D)
    update_mayoristas(D)
    update_turismo(D)
    update_merval(D)
    update_daily_series(D)
    apply_manual_overrides(D)  # datos preliminares de notas/consultoras hasta que llegue el oficial

    # Detectar nuevas publicaciones comparando contra el data.json previo
    pubs = detect_publicaciones(D_prev, D)
    nuevas_hoy = [p for p in pubs if p.get("detectado_at", "")[:10] == datetime.now(timezone.utc).strftime("%Y-%m-%d")]
    if nuevas_hoy:
        log(f"Publicaciones nuevas detectadas hoy: {len(nuevas_hoy)} ({', '.join(p['label']+' '+p['periodo'] for p in nuevas_hoy[:5])})", "ok")

    # Metadata
    D["_meta"] = {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "scraper_version": "1.1",
        "sources": ["INDEC", "BCRA", "OEDE", "Sec.Hacienda", "ArgentinaDatos", "Bluelytics", "Yahoo Finance"],
        "publicaciones": pubs[:50],  # Últimas 50 publicaciones detectadas
        "proximas": compute_proximas_publicaciones(D, days_ahead=7),
    }

    if args.dry_run:
        log("DRY RUN — no escribo data.json")
        log(f"Sumario: {len(D) - 1} series (excluye _meta)")
    else:
        OUT_PATH.write_text(json.dumps(D, ensure_ascii=False, indent=None, separators=(",", ":")),
                            encoding="utf-8")
        size_kb = OUT_PATH.stat().st_size / 1024
        log(f"OK — data.json escrito ({size_kb:.1f} KB)", "ok")

    return 0

if __name__ == "__main__":
    sys.exit(main())
