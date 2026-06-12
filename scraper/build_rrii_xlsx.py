"""Genera rrii_brutas.xlsx: reservas brutas diarias + factores de variación BCRA.

Replica el formato del informe manual "RRII Brutas al DD de mes.xlsx":
  Fecha | RRII Brutas | Var diaria RRII | Compras de divisas | Org. int. |
  Otras op. sector público | Efectivo mínimo | Otros | Var mensual % |
  Var i.a. % | Acum. anual RRII | Acum. anual compras
más un bloque resumen al pie (día / mensual / anual / por año / era Milei).

Fuente: API BCRA v4.0 monetarias
  var 1  = Reservas internacionales (saldo, USD M)
  var 78 = Variación por compra de divisas
  var 79 = Variación por organismos internacionales
  var 80 = Variación por otras operaciones del sector público
  var 81 = Variación por efectivo mínimo
  var 82 = Variación por otras operaciones

Incremental: si el XLSX ya existe, sólo trae fechas posteriores a la última.
"""
from __future__ import annotations
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import requests
import urllib3
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HERE = Path(__file__).parent
OUT = HERE.parent / "rrii_brutas.xlsx"
BCRA_API = "https://api.bcra.gob.ar/estadisticas/v4.0/monetarias/"
INICIO_HISTORICO = date(2004, 11, 16)
ERA_MILEI_DESDE = date(2023, 12, 11)  # primer día hábil de gestión

VARS = {
    "brutas": 1,
    "compras": 78,
    "org_int": 79,
    "sec_pub": 80,
    "efec_min": 81,
    "otros": 82,
}
HEADERS_XLSX = [
    "Fecha", "RRII Brutas", "Var diaria de RRII",
    "Var diaria de Compras de divisas", "Var diaria de Org. int.",
    "Var diaria de otras operaciones con el sector público",
    "Var diaria de efect. minimo", "Var diaria de otros",
    "Var mensual de RRII brutas", "Var i.a. de RRII brutas",
    "Variación absoluta acumulada anual RRII",
    "Variación absoluta acumulada anual Compra de Divisas",
]


def fetch_var(var_id: int, desde: date, hasta: date) -> dict[str, float]:
    """Trae una variable BCRA por rangos anuales. Devuelve {iso_date: valor}."""
    out: dict[str, float] = {}
    d0 = desde
    while d0 <= hasta:
        d1 = min(date(d0.year, 12, 31), hasta)
        for attempt in range(3):
            try:
                r = requests.get(
                    f"{BCRA_API}{var_id}",
                    params={"desde": d0.isoformat(), "hasta": d1.isoformat(), "limit": 3000},
                    timeout=40, verify=False,
                )
                r.raise_for_status()
                det = (r.json().get("results") or [{}])[0].get("detalle", [])
                for row in det:
                    if row.get("fecha") and row.get("valor") is not None:
                        out[row["fecha"]] = float(row["valor"])
                break
            except Exception as e:
                if attempt == 2:
                    print(f"WARN var {var_id} {d0}..{d1}: {e}")
                time.sleep(2 ** attempt)
        d0 = date(d0.year + 1, 1, 1)
    return out


def load_existing() -> list[dict]:
    """Lee las filas diarias del XLSX existente (ignora el bloque resumen)."""
    if not OUT.exists():
        return []
    try:
        wb = openpyxl.load_workbook(OUT, data_only=True)
        ws = wb.active
        rows = []
        for r in range(2, ws.max_row + 1):
            f = ws.cell(r, 1).value
            if not isinstance(f, datetime):
                continue  # bloque resumen / filas vacías
            rows.append({
                "fecha": f.date(),
                "brutas": ws.cell(r, 2).value,
                "compras": ws.cell(r, 4).value,
                "org_int": ws.cell(r, 5).value,
                "sec_pub": ws.cell(r, 6).value,
                "efec_min": ws.cell(r, 7).value,
                "otros": ws.cell(r, 8).value,
            })
        rows.sort(key=lambda x: x["fecha"])
        return rows
    except Exception as e:
        print(f"WARN no pude leer {OUT.name} existente ({e}), regenero completo")
        return []


def main() -> int:
    hoy = date.today()
    rows = load_existing()
    if rows:
        # Incremental: refrescar los últimos 10 días hábiles (factores llegan con rezago)
        desde = rows[-1]["fecha"] - timedelta(days=14)
        rows = [r for r in rows if r["fecha"] < desde]
        print(f"Incremental desde {desde} (filas previas conservadas: {len(rows)})")
    else:
        desde = INICIO_HISTORICO
        print(f"Generación completa desde {desde} (~{(hoy-desde).days} días, puede tardar)")

    series = {}
    for key, var_id in VARS.items():
        series[key] = fetch_var(var_id, desde, hoy)
        print(f"  var {var_id} ({key}): {len(series[key])} registros")

    fechas_nuevas = sorted(set(series["brutas"].keys()))
    for f_iso in fechas_nuevas:
        f = date.fromisoformat(f_iso)
        rows.append({
            "fecha": f,
            "brutas": series["brutas"].get(f_iso),
            "compras": series["compras"].get(f_iso),
            "org_int": series["org_int"].get(f_iso),
            "sec_pub": series["sec_pub"].get(f_iso),
            "efec_min": series["efec_min"].get(f_iso),
            "otros": series["otros"].get(f_iso),
        })
    rows.sort(key=lambda x: x["fecha"])
    if not rows:
        print("Sin datos — no escribo XLSX")
        return 1

    # Índices auxiliares para variaciones
    by_date = {r["fecha"]: r for r in rows}
    fechas = [r["fecha"] for r in rows]

    def busca_anterior(f: date) -> dict | None:
        """Último registro estrictamente anterior a f (para fechas no hábiles)."""
        import bisect
        i = bisect.bisect_left(fechas, f)
        return rows[i - 1] if i > 0 else None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RRII Brutas"
    bold = Font(bold=True)
    for c, h in enumerate(HEADERS_XLSX, 1):
        cell = ws.cell(1, c, h)
        cell.font = bold

    cierre_anio: dict[int, float] = {}   # brutas al último día de cada año
    acum_compras: dict[int, float] = {}  # suma de compras por año

    prev = None
    for r in rows:
        f = r["fecha"]
        b = r["brutas"]
        # acumuladores anuales
        if b is not None:
            cierre_anio[f.year] = b
        if r["compras"] is not None:
            acum_compras[f.year] = acum_compras.get(f.year, 0.0) + r["compras"]

        var_d = (b - prev["brutas"]) if (b is not None and prev and prev["brutas"] is not None) else None
        # var mensual %: vs último registro <= mismo día mes anterior
        var_m = None
        var_ia = None
        if b is not None:
            try:
                f_m = f.replace(month=f.month - 1) if f.month > 1 else f.replace(year=f.year - 1, month=12)
            except ValueError:
                f_m = (f.replace(day=1) - timedelta(days=1))
            ref_m = by_date.get(f_m) or busca_anterior(f_m)
            if ref_m and ref_m["brutas"]:
                var_m = b / ref_m["brutas"] - 1
            try:
                f_y = f.replace(year=f.year - 1)
            except ValueError:
                f_y = f.replace(year=f.year - 1, day=28)
            ref_y = by_date.get(f_y) or busca_anterior(f_y)
            if ref_y and ref_y["brutas"]:
                var_ia = b / ref_y["brutas"] - 1
        cierre_prev = cierre_anio.get(f.year - 1)
        acum_rrii = (b - cierre_prev) if (b is not None and cierre_prev is not None) else None
        acum_cmp = acum_compras.get(f.year)

        ws.append([
            datetime(f.year, f.month, f.day), b, round(var_d) if var_d is not None else None,
            r["compras"], r["org_int"], r["sec_pub"], r["efec_min"], r["otros"],
            round(var_m, 3) if var_m is not None else None,
            round(var_ia, 3) if var_ia is not None else None,
            round(acum_rrii) if acum_rrii is not None else None,
            round(acum_cmp) if acum_cmp is not None else None,
        ])
        prev = r

    # ── Bloque resumen ────────────────────────────────────────────────────────
    ws.append([])
    ult = rows[-1]
    b_ult = ult["brutas"]

    def delta_desde(f0: date) -> float | None:
        ref = by_date.get(f0) or busca_anterior(f0)
        return (b_ult - ref["brutas"]) if (ref and ref["brutas"] is not None and b_ult is not None) else None

    fila_resumen = [
        ("Día", (b_ult - busca_anterior(ult["fecha"])["brutas"]) if len(rows) > 1 else None),
        ("Mensual", delta_desde(ult["fecha"] - timedelta(days=30))),
        ("Anual", delta_desde(date(ult["fecha"].year - 1, 12, 31))),
        ("Era Milei", delta_desde(ERA_MILEI_DESDE)),
    ]
    for lbl, v in fila_resumen:
        ws.append([lbl, round(v) if v is not None else None])
        ws.cell(ws.max_row, 1).font = bold

    # Por año: delta de brutas + suma de factores
    ws.append([])
    ws.append([None, "Año", "Δ RRII", "Compras", "Org. int.", "Sector púb.", "Efect. mín.", "Otros"])
    ws.cell(ws.max_row, 2).font = bold
    for anio in sorted(set(f.year for f in fechas)):
        if anio < ult["fecha"].year - 2:
            continue  # solo últimos 3 años en el resumen
        cierre_ant = cierre_anio.get(anio - 1)
        delta = (cierre_anio[anio] - cierre_ant) if (anio in cierre_anio and cierre_ant is not None) else None
        sums = {}
        for k in ("compras", "org_int", "sec_pub", "efec_min", "otros"):
            vals = [r[k] for r in rows if r["fecha"].year == anio and r[k] is not None]
            sums[k] = round(sum(vals)) if vals else None
        ws.append([None, anio, round(delta) if delta is not None else None,
                   sums["compras"], sums["org_int"], sums["sec_pub"], sums["efec_min"], sums["otros"]])

    # Anchos de columna
    widths = [11, 11, 10, 12, 10, 14, 12, 10, 9, 9, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(OUT)
    n_diarias = len(rows)
    print(f"OK {OUT.name}: {n_diarias} filas diarias (último {ult['fecha']}, brutas USD {b_ult:,.0f}M)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
