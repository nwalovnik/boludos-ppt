"""
Carga el Excel manual de Google Forms en historical_despidos.json.

Reemplaza los eventos con fuente!='auto-rss/auto-link' (manuales) por los del Excel.
Los eventos automaticos se preservan.

Uso:
    python load_manual_excel.py <ruta_excel>      # archivo local
    python load_manual_excel.py --gsheet <URL>    # baja CSV del Google Sheet publico
    python load_manual_excel.py                   # usa default en Downloads
    python load_manual_excel.py --gsheet-env      # toma URL de env var GSHEET_URL
"""
from __future__ import annotations

import io
import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import requests
from openpyxl import load_workbook

SCRAPER_DIR = Path(__file__).resolve().parent
BEDROCK = SCRAPER_DIR / "historical_despidos.json"
DEFAULT_XLSX = Path.home() / "Downloads" / "CONTADOR DE CIERRE DE EMPRESAS Y DESPIDOS (respuestas) (1).xlsx"


def detectar_tipo(texto: str) -> str:
    t = (texto or "").lower()
    KW_CIERRE = ["cierre", "cerro", "cerró", "cerraron", "quiebra", "quebró",
                 "concurso de acreedores", "liquidacion", "liquidación",
                 "deja de producir", "cierre de planta"]
    KW_SUSP = ["suspension", "suspensión", "suspensiones", "suspendido", "suspendidos"]
    if any(k in t for k in KW_CIERRE):
        return "cierre"
    if any(k in t for k in KW_SUSP):
        return "suspensión"
    return "despido"


def gsheet_csv_url(sheet_url: str) -> str:
    """Convierte una URL de Google Sheet en su URL de export CSV (gid=0)."""
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", sheet_url)
    if not m:
        raise ValueError(f"URL de Google Sheet invalida: {sheet_url}")
    sheet_id = m.group(1)
    gid_m = re.search(r"[#&?]gid=(\d+)", sheet_url)
    gid = gid_m.group(1) if gid_m else "0"
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


def cargar_csv(csv_text: str) -> list[dict]:
    """Parsea un CSV de Google Sheet con la misma estructura del Form."""
    import csv as csv_mod
    rows = list(csv_mod.reader(io.StringIO(csv_text)))
    if not rows:
        return []
    headers = [h.strip() for h in rows[0]]
    return _filas_a_eventos(headers, rows[1:])


def _filas_a_eventos(headers, rows):
    col = {h: i for i, h in enumerate(headers)}

    def get(row, name, default=""):
        for h, i in col.items():
            if name.lower() in h.lower() and i < len(row):
                v = row[i]
                return v if v not in (None, "") else default
        return default

    eventos = []
    for row in rows:
        if not any(row):
            continue
        fecha_raw = get(row, "Fecha")
        if isinstance(fecha_raw, datetime):
            fecha_str = fecha_raw.strftime("%Y-%m-%d")
        elif isinstance(fecha_raw, date):
            fecha_str = fecha_raw.isoformat()
        else:
            s = str(fecha_raw or "").strip()[:10]
            # Soportar DD/MM/YYYY (formato comun de Google Sheets)
            m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
            if m:
                d, mo, y = m.groups()
                fecha_str = f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
            else:
                fecha_str = s
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", fecha_str):
            continue

        try:
            emp_raw = str(get(row, "Cantidad de empleados", 1)).replace(",", ".").replace(".", "")
            emp = int(float(emp_raw)) if emp_raw else 1
        except (ValueError, TypeError):
            emp = 1

        cerro_raw = str(get(row, "cerr") or "").lower().strip()
        cerro = cerro_raw in ("si", "sí", "yes", "1", "true")

        coment = str(get(row, "Aclaraciones") or "")
        estado = str(get(row, "ESTADO") or "")
        rubro_pri = str(get(row, "Rubro Principal") or "").strip()
        rubro_otro = str(get(row, "Rubro Principal-si") or "").strip()
        rubro = rubro_pri if rubro_pri and rubro_pri.lower() != "otro" else (rubro_otro or rubro_pri or "Otro")

        eventos.append({
            "fecha":      fecha_str,
            "empresa":    str(get(row, "Empresa") or "").strip(),
            "rubro":      rubro,
            "empleados":  emp,
            "comentario": coment,
            "provincia":  str(get(row, "Provincia") or "").strip(),
            "municipio":  str(get(row, "municipio") or "").strip(),
            "estado":     estado,
            "sector":     str(get(row, "SECTOR") or "").strip(),
            "cerro":      cerro,
            "tipo":       detectar_tipo(coment + " " + estado),
            "fuente":     "manual",
            "url":        "",
        })
    return eventos


def cargar(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [(str(h).strip() if h else "") for h in rows[0]]
    eventos = _filas_a_eventos(headers, rows[1:])
    wb.close()
    return eventos


def cargar_desde_gsheet(sheet_url: str) -> list[dict]:
    csv_url = gsheet_csv_url(sheet_url)
    print(f"  Bajando CSV: {csv_url}")
    r = requests.get(csv_url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
    r.raise_for_status()
    return cargar_csv(r.text)


def main(argv: list[str] | None = None) -> int:
    argv = argv or sys.argv[1:]

    if argv and argv[0] == "--gsheet" and len(argv) > 1:
        sheet_url = argv[1]
        print(f"Cargando Google Sheet: {sheet_url}")
        eventos_manual = cargar_desde_gsheet(sheet_url)
    elif argv and argv[0] == "--gsheet-env":
        sheet_url = os.environ.get("GSHEET_URL", "").strip()
        if not sheet_url:
            print("WARN: GSHEET_URL no esta seteada, salgo sin tocar el bedrock")
            return 0
        print(f"Cargando Google Sheet (env): {sheet_url}")
        eventos_manual = cargar_desde_gsheet(sheet_url)
    else:
        xlsx_path = Path(argv[0]) if argv else DEFAULT_XLSX
        if not xlsx_path.exists():
            print(f"ERROR: Excel no encontrado: {xlsx_path}")
            return 1
        print(f"Cargando Excel: {xlsx_path}")
        eventos_manual = cargar(xlsx_path)

    print(f"  Eventos manuales leidos: {len(eventos_manual)}")

    if not BEDROCK.exists():
        # Bedrock vacio: arranco solo con los manuales
        bedrock = {"eventos": eventos_manual}
    else:
        bedrock = json.load(open(BEDROCK, "r", encoding="utf-8"))
        AUTO_FUENTES = {"auto-rss", "auto-link"}
        auto = [e for e in bedrock.get("eventos", []) if e.get("fuente") in AUTO_FUENTES]
        print(f"  Eventos automaticos preservados: {len(auto)}")
        bedrock["eventos"] = eventos_manual + auto

    with open(BEDROCK, "w", encoding="utf-8") as f:
        json.dump(bedrock, f, ensure_ascii=False, indent=None)
    print(f"  Bedrock actualizado: {len(bedrock['eventos'])} eventos -> {BEDROCK}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
